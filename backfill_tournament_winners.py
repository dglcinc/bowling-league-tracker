"""
Backfill 2nd and 3rd place tournament winners from historical XLS files.

For each non-COVID season:
  - Re-reads the Payout Formula sheet to extract all 3 places
  - Deletes existing TournamentEntry rows for those tournament weeks
  - Re-inserts all 3 places with game1 = 300 / 200 / 100 (sort order)

Usage (Flask app must be stopped first):
    python3 backfill_tournament_winners.py [--dry-run] [season_name]
"""

import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl

DRY_RUN    = '--dry-run' in sys.argv
SEASON_ARG = next((a for a in sys.argv[1:] if not a.startswith('--')), None)

SPREADSHEET_DIR = os.path.expanduser('~/OneDrive - DGLC/Claude/Historic Scoresheets/')

# Must match seed_historical_seasons.py SEASONS list
SEASONS = [
    {'filename': 'scoring 2017-2018 - week 23.xlsx', 'name': '2017-2018', 'num_weeks': 22},
    {'filename': 'scoring 2018-2019 - week 23.xlsx', 'name': '2018-2019', 'num_weeks': 22},
    # 2019-2020 is covid_season — no tournaments
    {'filename': 'scoring 2021-2022 - Week 23.xlsx', 'name': '2021-2022', 'num_weeks': 22},
    {'filename': 'scoring 2022-2023 - Week 23.xlsx', 'name': '2022-2023', 'num_weeks': 22},
    {'filename': 'scoring 2023-2024 - Week 23.xlsx', 'name': '2023-2024', 'num_weeks': 22},
    {'filename': 'scoring 2024-2025 - Week 23.xlsx', 'name': '2024-2025', 'num_weeks': 22},
]

PLACE_SCORES = {1: 300, 2: 200, 3: 100}


def read_payout_winners(wb):
    """From Payout Formula sheet extract {tt: {place: name}}."""
    if 'Payout Formula' not in wb.sheetnames:
        return {}
    ws = wb['Payout Formula']
    rows = list(ws.iter_rows(values_only=True))

    results = {}
    current_key = None
    in_tournaments = False

    for row in rows:
        if not any(v is not None for v in row):
            continue
        if row[0] == 'Tournaments':
            in_tournaments = True
            continue
        if row[0] in ('Sub-Total', 'Weekly Prizes', 'Team Play'):
            in_tournaments = False
            continue
        if not in_tournaments:
            continue

        if row[1] is not None and row[2] is None:
            label = str(row[1]).strip().lower()
            if 'bedford' in label or 'belyea' in label:
                current_key = 'indiv_hcp_1'
            elif 'rose' in label or 'chad' in label:
                current_key = 'indiv_hcp_2'
            elif 'club' in label:
                current_key = 'indiv_scratch'
            else:
                current_key = None
            if current_key and current_key not in results:
                results[current_key] = {}

        elif current_key and row[2] is not None and row[5] is not None:
            place_str = str(row[2]).strip().lower()
            winner = str(row[6]).strip() if row[6] else ''
            # Row[6] might be a dash (—) meaning nobody recorded
            if not winner or winner in ('—', '-', '–'):
                continue
            if '1st' in place_str:
                results[current_key][1] = winner
            elif '2nd' in place_str:
                results[current_key][2] = winner
            elif '3rd' in place_str:
                results[current_key][3] = winner

    return results


def find_bowler(raw_name, Bowler):
    """Try to match raw_name to a Bowler row. Handles several formats:
    - 'Last'                 → match last_name
    - 'Last, First'          → match last_name (part before comma)
    - 'First Last'           → try last word as last_name
    - 'LastK' (trailing cap) → try stripping trailing uppercase letter(s)
    - 'Last, Last2 split'    → take first comma-segment only
    """
    if not raw_name:
        return None, raw_name
    raw_name = raw_name.strip().strip('—–-').strip()
    if not raw_name:
        return None, None

    candidates = []

    if ',' in raw_name:
        # "Last, First" or "Last, Last2 split" → use part before first comma
        candidates.append(raw_name.split(',')[0].strip())
    else:
        words = raw_name.split()
        # "First Last" format: try last word
        if len(words) > 1:
            candidates.append(words[-1])
        # Single word or fallback: use as-is
        candidates.append(words[0])

    # Also try stripping a trailing uppercase letter (e.g. "FaehnerK" → "Faehner")
    extra = []
    for c in candidates:
        if c and c[-1].isupper() and len(c) > 1 and c[-2].islower():
            extra.append(c[:-1])
    candidates.extend(extra)

    for last in candidates:
        bowler = Bowler.query.filter(
            Bowler.last_name.ilike(last)
        ).first()
        if bowler:
            return bowler, None

    return None, raw_name


def backfill(app):
    from models import db, Season, Week, Bowler, TournamentEntry

    with app.app_context():
        for cfg in SEASONS:
            if SEASON_ARG and cfg['name'] != SEASON_ARG:
                continue

            filepath = os.path.join(SPREADSHEET_DIR, cfg['filename'])
            if not os.path.exists(filepath):
                print(f'[{cfg["name"]}] SKIP — file not found: {filepath}')
                continue

            season = Season.query.filter_by(name=cfg['name']).first()
            if not season:
                print(f'[{cfg["name"]}] SKIP — season not in DB')
                continue

            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            winners = read_payout_winners(wb)

            if not winners:
                print(f'[{cfg["name"]}] No payout winners found in spreadsheet')
                continue

            num_weeks = cfg['num_weeks']
            tt_to_wk = {
                'indiv_scratch': num_weeks + 2,
                'indiv_hcp_1':   num_weeks + 3,
                'indiv_hcp_2':   num_weeks + 4,
            }

            print(f'\n[{cfg["name"]}]')
            for tt, places in winners.items():
                wk_num = tt_to_wk.get(tt)
                if not wk_num:
                    continue

                # Ensure the Week row exists and is_entered
                wk = Week.query.filter_by(season_id=season.id, week_num=wk_num).first()
                if not wk:
                    print(f'  {tt}: week {wk_num} not found — skipping')
                    continue

                # Delete existing entries for this tournament week
                existing = TournamentEntry.query.filter_by(
                    season_id=season.id, week_num=wk_num
                ).all()
                print(f'  {tt} (week {wk_num}): removing {len(existing)} existing entries')
                if not DRY_RUN:
                    for e in existing:
                        db.session.delete(e)
                    db.session.flush()

                # Insert all 3 places
                for place in [1, 2, 3]:
                    raw_name = places.get(place)
                    if not raw_name:
                        print(f'    Place {place}: no name recorded')
                        continue

                    bowler, guest = find_bowler(raw_name, Bowler)
                    score = PLACE_SCORES[place]

                    display = bowler.last_name if bowler else f'guest:{guest}'
                    print(f'    Place {place}: {raw_name!r} → {display}  (game1={score})')

                    if not DRY_RUN:
                        te = TournamentEntry(
                            season_id=season.id,
                            week_num=wk_num,
                            bowler_id=bowler.id if bowler else None,
                            guest_name=guest,
                            handicap=0,
                            game1=score,
                        )
                        db.session.add(te)

                if not DRY_RUN:
                    wk.is_entered = True

            if not DRY_RUN:
                db.session.commit()
                print(f'  Committed.')


def main():
    from app import create_app
    app = create_app()

    if DRY_RUN:
        print('DRY RUN — no DB writes.')

    backfill(app)


if __name__ == '__main__':
    main()

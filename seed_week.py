#!/usr/bin/env python3
"""
Import one week's scores into the 2025-2026 historical season.

Run on Mac with Flask app stopped:
  python seed_week.py <week_num> "/path/to/scoring 2025-2026 - Week 22.xlsx"

Examples:
  python seed_week.py 1  "/Users/you/OneDrive - DGLC/Claude/scoring 2025-2026 - Week 22.xlsx"
  python seed_week.py 11 "/Users/you/OneDrive - DGLC/Claude/scoring 2025-2026 - Week 22.xlsx"

What this does for each week:
  1. Loads actual team points from the 'team scoring' sheet (ground truth).
  2. Verifies the scheduled lane assignment by checking which competing-team pairs
     sum to exactly 8 total points.  Mismatches flag weeks where the printed
     schedule differed from what actually happened on the lanes.
  3. If the verification finds a different assignment (unambiguous), it updates
     the ScheduleEntry in the DB for that week and reports the change.
  4. Loads individual bowler game scores from their personal sheets.
  5. Creates MatchupEntry records (matchup_num = team number, for individual stats).
  6. Creates TeamPoints from the spreadsheet — these are the authoritative points
     for standings, not recomputed from the imported scores.
  7. Marks the week as entered and saves a JSON snapshot.

Week 22 (position night not yet bowled) is skipped — enter it live through the app.
"""

import sys
import os
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from app import create_app
from models import db, Season, Team, Bowler, Roster, Week, ScheduleEntry, MatchupEntry, TeamPoints
from snapshots import save_snapshot
from config import Config

SEASON_NAME = "2025-2026"

NON_BOWLER_SHEETS = {
    'Instructions', 'Parameters', '2025 Banquet', 'wkly alpha', 'YTD alpha',
    'wkly high average', 'High Games ', 'team scoring', 'dummy', 'blinds',
    'Payout Formula', 'indiv payout', 'final handicap',
}

LANE_PAIRS = ["17/19", "18/20", "21/23", "22/24"]


# ─── spreadsheet readers ──────────────────────────────────────────────────────

def load_bowler_week(wb, sheet_name, week_num):
    """
    Return (g1, g2, g3) for a bowler's games in week_num, or None if absent.
    """
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    week_row  = rows[6]
    game1_row = rows[8]
    game2_row = rows[9]
    game3_row = rows[10]

    col = None
    for c in range(2, 25):
        if c < len(week_row) and week_row[c] == week_num:
            col = c
            break
    if col is None:
        return None

    g1 = game1_row[col] if col < len(game1_row) else None
    g2 = game2_row[col] if col < len(game2_row) else None
    g3 = game3_row[col] if col < len(game3_row) else None

    if g1 is None and g2 is None and g3 is None:
        return None
    if g1 == 0 and g2 == 0 and g3 == 0:
        return None

    return (
        int(g1) if g1 is not None else None,
        int(g2) if g2 is not None else None,
        int(g3) if g3 is not None else None,
    )


def load_team_scoring_week(wb, week_num):
    """
    Return dict: {team_num: {'A': pts, 'B': pts, 'total': pts}} for the given week.
    Columns: week, date, T1A, T1B, T1tot, T2A, T2B, T2tot, T3A, T3B, T3tot, T4A, T4B, T4tot, grand
    """
    ws = wb['team scoring']
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[7:]:
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue
        if int(row[0]) != week_num:
            continue

        def p(v):
            return float(v) if v is not None else 0.0

        return {
            1: {'A': p(row[2]),  'B': p(row[3]),  'total': p(row[4])},
            2: {'A': p(row[5]),  'B': p(row[6]),  'total': p(row[7])},
            3: {'A': p(row[8]),  'B': p(row[9]),  'total': p(row[10])},
            4: {'A': p(row[11]), 'B': p(row[12]), 'total': p(row[13])},
        }
    return None


# ─── lane assignment verification ────────────────────────────────────────────

def detect_competition_structure(team_totals):
    """
    Given {team_num: total_pts}, determine which teams competed against each other.

    In every regular week the two teams in each competition share exactly 8 pts
    between them (or 0 if the week was cancelled).  There are only three ways to
    pair four teams:
        {1v2, 3v4}   {1v3, 2v4}   {1v4, 2v3}

    Returns (structure, ambiguous) where structure is e.g. [(1,2),(3,4)].
    ambiguous=True when the totals don't uniquely identify one pairing
    (happens when all pts are 0, or when multiple pairings could explain the totals).
    """
    all_structures = [
        [(1, 2), (3, 4)],
        [(1, 3), (2, 4)],
        [(1, 4), (2, 3)],
    ]

    grand_total = sum(team_totals.values())

    # Cancelled / no-bowl week — can't determine
    if grand_total == 0:
        return None, True

    matches = []
    for structure in all_structures:
        (a, b), (c, d) = structure
        sum1 = team_totals[a] + team_totals[b]
        sum2 = team_totals[c] + team_totals[d]
        if abs(sum1 - 8) < 0.01 and abs(sum2 - 8) < 0.01:
            matches.append(structure)

    if len(matches) == 1:
        return matches[0], False
    return None, True


def structure_from_schedule(sched_entries, team_num_for_id):
    """
    Derive competition structure [(t1,t2),(t3,t4)] from ScheduleEntry records.
    Matchup_nums 1&2 share pairing A; matchup_nums 3&4 share pairing B.
    """
    by_mnum = {s.matchup_num: s for s in sched_entries}
    if 1 not in by_mnum or 3 not in by_mnum:
        return None
    a = (team_num_for_id[by_mnum[1].team1_id], team_num_for_id[by_mnum[1].team2_id])
    b = (team_num_for_id[by_mnum[3].team1_id], team_num_for_id[by_mnum[3].team2_id])
    # Normalise so the lower team number comes first in each pair
    return [tuple(sorted(a)), tuple(sorted(b))]


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Usage: python seed_week.py <week_num> <xlsx_path>")
        sys.exit(1)

    week_num  = int(sys.argv[1])
    xlsx_path = sys.argv[2]

    if week_num == 22:
        print("Week 22 is the position night not yet bowled. Enter it live through the app.")
        sys.exit(0)

    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  Week {week_num}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    app = create_app()
    with app.app_context():

        season = Season.query.filter_by(name=SEASON_NAME).first()
        if not season:
            print(f"ERROR: Season '{SEASON_NAME}' not found. Run seed_historical.py first.")
            sys.exit(1)

        week = Week.query.filter_by(season_id=season.id, week_num=week_num).first()
        if not week:
            print(f"ERROR: Week {week_num} not found. Run seed_historical.py first.")
            sys.exit(1)

        if week.is_position_night:
            print(f"Week {week_num} is a position night. Enter it live through the app.")
            sys.exit(0)

        teams_db      = {t.number: t for t in Team.query.filter_by(season_id=season.id).all()}
        team_num_for_id = {t.id: t.number for t in teams_db.values()}

        # ── Step 1: load ground-truth team points from spreadsheet ─────────
        actual = load_team_scoring_week(wb, week_num)
        if actual is None:
            print(f"ERROR: No team scoring data found for week {week_num}.")
            sys.exit(1)

        print(f"\nSpreadsheet team points (from 'team scoring' sheet):")
        for tnum in [1, 2, 3, 4]:
            d = actual[tnum]
            print(f"  Team {tnum} ({teams_db[tnum].name}):  "
                  f"A={d['A']:.1f}  B={d['B']:.1f}  Total={d['total']:.1f}")

        team_totals = {t: actual[t]['total'] for t in [1, 2, 3, 4]}

        # ── Step 2: detect actual competition structure ────────────────────
        detected_structure, ambiguous = detect_competition_structure(team_totals)

        sched_entries = (ScheduleEntry.query
                         .filter_by(season_id=season.id, week_num=week_num)
                         .order_by(ScheduleEntry.matchup_num).all())
        scheduled_structure = structure_from_schedule(sched_entries, team_num_for_id)

        print(f"\nSchedule says:  {scheduled_structure}")

        if ambiguous:
            print(f"Detected:       AMBIGUOUS (week may be cancelled or totals are symmetrical)")
            print(f"  → Using scheduled assignment.")
            final_structure = scheduled_structure
        elif detected_structure:
            # Normalise for comparison
            det_norm = [tuple(sorted(p)) for p in detected_structure]
            sch_norm = [tuple(sorted(p)) for p in (scheduled_structure or [])]
            # Both pairs need to match (order of pairs doesn't matter)
            if set(map(tuple, det_norm)) == set(map(tuple, sch_norm)):
                print(f"Detected:       {detected_structure}  ✓ MATCHES schedule")
                final_structure = scheduled_structure
            else:
                print(f"Detected:       {detected_structure}  ⚠ DIFFERS from schedule")
                print(f"  → Updating ScheduleEntry for week {week_num} to match actual.")
                final_structure = detected_structure

                # Rebuild schedule entries for this week
                ScheduleEntry.query.filter_by(
                    season_id=season.id, week_num=week_num
                ).delete()
                (a1, a2), (b1, b2) = final_structure
                for mnum, t1, t2, lp in [
                    (1, a1, a2, LANE_PAIRS[0]),
                    (2, a1, a2, LANE_PAIRS[1]),
                    (3, b1, b2, LANE_PAIRS[2]),
                    (4, b1, b2, LANE_PAIRS[3]),
                ]:
                    db.session.add(ScheduleEntry(
                        season_id=season.id, week_num=week_num, matchup_num=mnum,
                        team1_id=teams_db[t1].id, team2_id=teams_db[t2].id, lane_pair=lp,
                    ))
                db.session.flush()
        else:
            final_structure = scheduled_structure

        # ── Step 3: load bowler scores for this week ──────────────────────
        # Maps team_num -> [(Bowler, (g1,g2,g3))]
        team_bowlers = {1: [], 2: [], 3: [], 4: []}

        bowlers_db = {b.last_name: b for b in Bowler.query.all()}
        roster_map = {}  # bowler_id -> team_num
        for r in Roster.query.filter_by(season_id=season.id, active=True).all():
            roster_map[r.bowler_id] = team_num_for_id.get(r.team_id)

        for sheet_name in wb.sheetnames:
            if sheet_name in NON_BOWLER_SHEETS:
                continue
            bowler = bowlers_db.get(sheet_name)
            if not bowler:
                continue
            games = load_bowler_week(wb, sheet_name, week_num)
            if games is None:
                continue
            tnum = roster_map.get(bowler.id)
            if tnum is None:
                continue
            team_bowlers[tnum].append((bowler, games))

        print(f"\nBowlers with scores this week:")
        total_bowlers = 0
        for tnum in [1, 2, 3, 4]:
            blist = team_bowlers[tnum]
            total_bowlers += len(blist)
            names = [b.last_name for b, _ in sorted(blist, key=lambda x: x[0].last_name)]
            print(f"  Team {tnum} ({teams_db[tnum].name}): {len(blist)} — {', '.join(names)}")

        # ── Step 4: clear and re-insert ───────────────────────────────────
        MatchupEntry.query.filter_by(season_id=season.id, week_num=week_num).delete()
        TeamPoints.query.filter_by(season_id=season.id, week_num=week_num).delete()
        db.session.flush()

        # ── Step 5: MatchupEntry — matchup_num = team number (for individual stats)
        # Both teams in a competition are stored at their own team-number matchup_num.
        # Individual stats (averages, handicaps) don't use matchup_num, so this is fine.
        # TeamPoints come from the spreadsheet rather than score_matchup() recalculation.
        entry_count = 0
        for tnum in [1, 2, 3, 4]:
            for bowler, games in sorted(team_bowlers[tnum], key=lambda x: x[0].last_name):
                db.session.add(MatchupEntry(
                    season_id=season.id,
                    week_num=week_num,
                    matchup_num=tnum,            # simplified: matchup_num = team number
                    team_id=teams_db[tnum].id,
                    bowler_id=bowler.id,
                    is_blind=False,
                    lane_side='A',
                    game1=games[0],
                    game2=games[1],
                    game3=games[2],
                ))
                entry_count += 1

        db.session.flush()

        # ── Step 6: TeamPoints — from spreadsheet (authoritative historical record)
        # A-side points → matchup_num 1 (teams in group A) or 3 (group B)
        # B-side points → matchup_num 2 (group A) or 4 (group B)
        # Group A teams = the first pair in final_structure (use matchup_nums 1&2)
        # Group B teams = the second pair (use matchup_nums 3&4)
        if final_structure:
            group_a_teams = set(final_structure[0])
            group_b_teams = set(final_structure[1])
        else:
            group_a_teams = {1, 2}
            group_b_teams = {3, 4}

        tp_count = 0
        for tnum in [1, 2, 3, 4]:
            if tnum in group_a_teams:
                mnum_a, mnum_b = 1, 2
            else:
                mnum_a, mnum_b = 3, 4

            for mnum, pts in [(mnum_a, actual[tnum]['A']), (mnum_b, actual[tnum]['B'])]:
                db.session.add(TeamPoints(
                    season_id=season.id,
                    week_num=week_num,
                    matchup_num=mnum,
                    team_id=teams_db[tnum].id,
                    points_earned=pts,
                ))
                tp_count += 1

        # Mark week as entered
        week.is_entered = True
        db.session.commit()

        print(f"\nImported: {entry_count} bowler score entries, {tp_count} team point records")

        # ── Step 7: JSON snapshot ─────────────────────────────────────────
        try:
            save_snapshot(season.id, week_num, Config.SNAPSHOT_DIR)
            print(f"✓ JSON snapshot saved  →  {Config.SNAPSHOT_DIR}/2025-2026-wk{week_num:02d}.json")
        except Exception as e:
            print(f"⚠ Snapshot error: {e}")

        # ── Summary ───────────────────────────────────────────────────────
        print(f"\nWeek {week_num} complete.")
        print(f"  {total_bowlers} bowlers entered | "
              f"Team pts: "
              f"T1={actual[1]['total']:.1f}  "
              f"T2={actual[2]['total']:.1f}  "
              f"T3={actual[3]['total']:.1f}  "
              f"T4={actual[4]['total']:.1f}")


if __name__ == '__main__':
    main()

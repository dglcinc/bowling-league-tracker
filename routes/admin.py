"""
Admin routes: season setup, roster management, schedule entry, season rollover.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app
from models import db, Season, Team, Roster, Bowler, Week, ScheduleEntry, MatchupEntry, LeagueSettings, LinkedAccount, ViewerPermission, TournamentEntry, ClubChampionshipResult
from extensions import cache
from datetime import date, timedelta
import io

admin_bp = Blueprint('admin', __name__)

# Post-season tournament weeks added after every regular season.
# Keys are generic — display names are configured per season in the seasons table.
_POSTSEASON_WEEKS = [
    ('club_championship', True),   # Club team championship — scored as position night
    ('indiv_scratch',     False),  # Individual scratch championship
    ('indiv_hcp_1',       False),  # Individual handicap tournament 1
    ('indiv_hcp_2',       False),  # Individual handicap tournament 2
]


def _add_postseason_weeks(season, num_regular_weeks):
    """Append 4 post-season tournament weeks after the regular season."""
    for offset, (tt, is_pos) in enumerate(_POSTSEASON_WEEKS, start=1):
        wn = num_regular_weeks + offset
        wk = Week(
            season_id=season.id,
            week_num=wn,
            tournament_type=tt,
            is_position_night=is_pos,
        )
        if season.start_date:
            wk.date = season.start_date + timedelta(weeks=wn - 1)
        db.session.add(wk)


# ---------------------------------------------------------------------------
# Seasons
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons')
def seasons():
    all_seasons = Season.query.order_by(Season.name.desc()).all()
    return render_template('admin/seasons.html', seasons=all_seasons)


@admin_bp.route('/seasons/new', methods=['GET', 'POST'])
def new_season():
    if request.method == 'POST':
        name = request.form['name'].strip()
        start_date_str = request.form.get('start_date')
        num_weeks = int(request.form.get('num_weeks', 23))
        half_boundary = int(request.form.get('half_boundary_week', 11))

        if Season.query.filter_by(name=name).first():
            flash(f'Season "{name}" already exists.', 'danger')
            return redirect(url_for('admin.new_season'))

        # Deactivate current active season
        Season.query.filter_by(is_active=True).update({'is_active': False})

        bowling_format = request.form.get('bowling_format', 'single')

        # Default tournament display names from the most recent existing season
        prev = Season.query.order_by(Season.id.desc()).first()
        season = Season(
            name=name,
            num_weeks=num_weeks,
            half_boundary_week=half_boundary,
            is_active=True,
            bowling_format=bowling_format,
            name_club_championship=request.form.get('name_club_championship', '').strip() or (prev.name_club_championship if prev else 'Club Championship'),
            name_indiv_scratch=request.form.get('name_indiv_scratch', '').strip() or (prev.name_indiv_scratch if prev else 'Individual Scratch Championship'),
            name_indiv_hcp_1=request.form.get('name_indiv_hcp_1', '').strip() or (prev.name_indiv_hcp_1 if prev else 'Individual Handicap Tournament 1'),
            name_indiv_hcp_2=request.form.get('name_indiv_hcp_2', '').strip() or (prev.name_indiv_hcp_2 if prev else 'Individual Handicap Tournament 2'),
        )
        if start_date_str:
            season.start_date = date.fromisoformat(start_date_str)

        db.session.add(season)
        db.session.flush()  # get season.id

        # Create 4 default teams with generic names
        for i in range(1, 5):
            tname = request.form.get(f'team{i}_name', '').strip() or f'Team {i}'
            team = Team(season_id=season.id, number=i, name=tname)
            db.session.add(team)

        # Create regular week records
        for wn in range(1, num_weeks + 1):
            wk = Week(
                season_id=season.id,
                week_num=wn,
                is_position_night=(wn in [half_boundary, num_weeks - 1]),
            )
            if season.start_date:
                wk.date = season.start_date + timedelta(weeks=wn - 1)
            db.session.add(wk)

        # Create 4 post-season tournament weeks
        _add_postseason_weeks(season, num_weeks)

        db.session.commit()
        flash(f'Season "{name}" created. Add your roster and schedule next.', 'success')
        return redirect(url_for('admin.season_detail', season_id=season.id))

    return render_template('admin/new_season.html')


@admin_bp.route('/seasons/<int:season_id>')
def season_detail(season_id):
    from types import SimpleNamespace
    season = Season.query.get_or_404(season_id)
    teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all()
    roster_filter = request.args.get('roster_filter', 'active')
    roster_q = (Roster.query
                .filter_by(season_id=season_id)
                .join(Bowler)
                .order_by(Bowler.last_name))
    if roster_filter == 'active':
        roster_q = roster_q.filter(Roster.active == True)
    roster = roster_q.all()

    # In 'all' mode, append bowlers rostered in any prior season but not this
    # one — shown inline as inactive entries with an Add to Roster action.
    if roster_filter == 'all':
        season_bowler_ids = {r.bowler_id for r in roster}
        ever_rostered_ids = {r.bowler_id for r in
                             Roster.query.with_entities(Roster.bowler_id).distinct().all()}
        unrostered_ids = ever_rostered_ids - season_bowler_ids
        if unrostered_ids:
            for b in (Bowler.query
                      .filter(Bowler.id.in_(unrostered_ids))
                      .order_by(Bowler.last_name).all()):
                roster.append(SimpleNamespace(
                    id=None, bowler_id=b.id, bowler=b,
                    team=None, active=False,
                    prior_handicap=None, joined_week=None,
                ))
            roster.sort(key=lambda r: r.bowler.last_name.lower())

    # Build access map: bowler_id → most recent LinkedAccount (for Access column)
    accounts = LinkedAccount.query.filter(
        LinkedAccount.bowler_id.in_([r.bowler_id for r in roster])
    ).all()
    access_map = {}
    for a in accounts:
        existing = access_map.get(a.bowler_id)
        if not existing or (a.last_login and (not existing.last_login or a.last_login > existing.last_login)):
            access_map[a.bowler_id] = a
    sender_email = current_app.config.get('GRAPH_SENDER_EMAIL', '')
    graph_configured = bool(current_app.config.get('GRAPH_CLIENT_ID'))
    return render_template('admin/season_detail.html',
                           season=season, teams=teams, roster=roster,
                           roster_filter=roster_filter, access_map=access_map,
                           sender_email=sender_email,
                           graph_configured=graph_configured)


@admin_bp.route('/seasons/<int:season_id>/send-magic-links', methods=['POST'])
def send_magic_links(season_id):
    from routes.auth import send_otp_invite
    Season.query.get_or_404(season_id)
    bowler_ids = request.form.getlist('bowler_ids', type=int)
    if not bowler_ids:
        flash('No bowlers selected.', 'warning')
        return redirect(url_for('admin.season_detail', season_id=season_id))

    settings = db.session.get(LeagueSettings, 1)
    league_name = settings.league_name if settings else 'League Tracker'
    from models import _DEFAULT_INVITE_MESSAGE
    invite_body = (settings.invite_message if settings and settings.invite_message
                   else _DEFAULT_INVITE_MESSAGE)
    subject = f'{league_name}: invitation to the app'

    sent = failed = no_email = 0
    for bid in bowler_ids:
        bowler = Bowler.query.get(bid)
        if not bowler:
            continue
        if not bowler.email:
            no_email += 1
            continue
        ok, _ = send_otp_invite(bowler, subject=subject, invite_body=invite_body)
        if ok:
            sent += 1
        else:
            failed += 1

    parts = []
    if sent:
        parts.append(f'{sent} invite(s) sent')
    if no_email:
        parts.append(f'{no_email} skipped (no email on file)')
    if failed:
        parts.append(f'{failed} failed — check Graph API config')
    flash(', '.join(parts) + '.', 'success' if not failed else 'warning')
    return redirect(url_for('admin.season_detail', season_id=season_id))


@admin_bp.route('/seasons/<int:season_id>/send-email', methods=['POST'])
def send_email(season_id):
    import html as _html
    season = Season.query.get_or_404(season_id)
    subject = request.form.get('subject', '').strip()
    body_text = request.form.get('body', '').strip()
    recipient_mode = request.form.get('recipient_mode', 'selected')
    bcc_self = request.form.get('bcc_self') == '1'

    if not subject or not body_text:
        flash('Subject and body are required.', 'warning')
        return redirect(url_for('admin.season_detail', season_id=season_id))

    if recipient_mode == 'selected':
        bowler_ids = request.form.getlist('bowler_ids', type=int)
        bowlers = Bowler.query.filter(Bowler.id.in_(bowler_ids)).all() if bowler_ids else []
    elif recipient_mode == 'all_active':
        active_ids = [r.bowler_id for r in
                      Roster.query.filter_by(season_id=season_id, active=True).all()]
        bowlers = Bowler.query.filter(Bowler.id.in_(active_ids)).all()
    elif recipient_mode == 'team':
        team_id = request.form.get('team_id', type=int)
        if team_id:
            t_ids = [r.bowler_id for r in
                     Roster.query.filter_by(season_id=season_id, team_id=team_id, active=True).all()]
            bowlers = Bowler.query.filter(Bowler.id.in_(t_ids)).all()
        else:
            bowlers = []
    else:
        bowlers = []

    recipient_emails = [b.email for b in bowlers if b.email]
    if not recipient_emails:
        flash('No recipients with email addresses found.', 'warning')
        return redirect(url_for('admin.season_detail', season_id=season_id))

    html_body = '<p>' + _html.escape(body_text).replace('\n', '<br>') + '</p>'

    sender_email = current_app.config.get('GRAPH_SENDER_EMAIL', '')
    if bcc_self and sender_email:
        to_list = [sender_email]
        bcc_list = recipient_emails
    else:
        to_list = recipient_emails
        bcc_list = []

    try:
        _send_via_graph(current_app.config, subject, html_body, to_list, bcc_list)
        flash(f'Email sent to {len(recipient_emails)} recipient(s).', 'success')
    except Exception as exc:
        flash(f'Email failed: {exc}', 'danger')

    return redirect(url_for('admin.season_detail', season_id=season_id))


@admin_bp.route('/viewer-access', methods=['GET', 'POST'])
def viewer_access():
    if request.method == 'POST':
        enabled = set(request.form.getlist('enabled'))
        for perm in ViewerPermission.query.all():
            perm.viewer_accessible = perm.endpoint in enabled
        db.session.commit()
        flash('Viewer access updated.', 'success')
        return redirect(url_for('admin.viewer_access'))
    perms = ViewerPermission.query.order_by(ViewerPermission.label).all()
    return render_template('admin/viewer_access.html', perms=perms)


# ---------------------------------------------------------------------------
# Team editing
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/teams/<int:team_id>/edit', methods=['GET', 'POST'])
def edit_team(season_id, team_id):
    season = Season.query.get_or_404(season_id)
    team = Team.query.get_or_404(team_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if name:
            team.name = name
        team.captain_name = request.form.get('captain_name', '').strip() or None
        db.session.commit()
        flash(f'{team.name} updated.', 'success')
        return redirect(url_for('admin.season_detail', season_id=season_id))
    return render_template('admin/edit_team.html', season=season, team=team)


# ---------------------------------------------------------------------------
# Roster management
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/roster/add', methods=['GET', 'POST'])
def add_bowler(season_id):
    season = Season.query.get_or_404(season_id)
    teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all()
    # Bowlers not already on this season's roster
    existing_ids = [r.bowler_id for r in Roster.query.filter_by(season_id=season_id)]
    available = Bowler.query.filter(
        Bowler.id.notin_(existing_ids)
    ).order_by(Bowler.last_name).all() if existing_ids else Bowler.query.order_by(Bowler.last_name).all()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'existing':
            bowler_id = int(request.form['bowler_id'])
        else:
            # Create new bowler
            bowler = Bowler(
                last_name=request.form['last_name'].strip(),
                first_name=request.form.get('first_name', '').strip() or None,
                nickname=request.form.get('nickname', '').strip() or None,
                email=request.form.get('email', '').strip() or None,
            )
            db.session.add(bowler)
            db.session.flush()
            bowler_id = bowler.id

        team_id = int(request.form['team_id'])
        prior_hcp = int(request.form.get('prior_handicap') or 0)
        joined_week = int(request.form.get('joined_week') or 1)

        roster_entry = Roster(
            bowler_id=bowler_id,
            season_id=season_id,
            team_id=team_id,
            active=True,
            prior_handicap=prior_hcp,
            joined_week=joined_week,
        )
        db.session.add(roster_entry)
        db.session.commit()
        flash('Bowler added to roster.', 'success')
        return redirect(url_for('admin.season_detail', season_id=season_id))

    preselect_id = request.args.get('bowler_id', type=int)
    return render_template('admin/add_bowler.html',
                           season=season, teams=teams, available=available,
                           preselect_id=preselect_id)


@admin_bp.route('/settings', methods=['GET', 'POST'])
def league_settings():
    settings = db.session.get(LeagueSettings, 1)
    if not settings:
        settings = LeagueSettings(id=1)
        db.session.add(settings)
        db.session.commit()
    active_season = Season.query.filter_by(is_active=True).order_by(Season.id.desc()).first()
    if request.method == 'POST':
        league_name = request.form.get('league_name', '').strip()
        if league_name:
            settings.league_name = league_name
        settings.use_nickname = (request.form.get('use_nickname') == 'on')
        settings.show_captain_name = (request.form.get('show_captain_name') == 'on')
        invite_msg = request.form.get('invite_message', '').strip()
        if invite_msg:
            settings.invite_message = invite_msg
        if active_season:
            try:
                active_season.handicap_base = int(request.form.get('handicap_base', active_season.handicap_base))
                active_season.blind_scratch = int(request.form.get('blind_scratch', active_season.blind_scratch))
                active_season.blind_handicap = int(request.form.get('blind_handicap', active_season.blind_handicap))
            except (ValueError, TypeError):
                pass
            arrival = request.form.get('arrival_time', '').strip()
            start = request.form.get('start_time', '').strip()
            if arrival:
                active_season.arrival_time = arrival
            if start:
                active_season.start_time = start
        db.session.commit()
        flash('League settings saved.', 'success')
        return redirect(url_for('admin.seasons'))
    from models import _DEFAULT_INVITE_MESSAGE
    return render_template('admin/league_settings.html', settings=settings,
                           active_season=active_season,
                           default_invite_message=_DEFAULT_INVITE_MESSAGE)


@admin_bp.route('/bowlers/<int:bowler_id>/edit', methods=['GET', 'POST'])
def edit_bowler(bowler_id):
    bowler = Bowler.query.get_or_404(bowler_id)
    season_id = request.args.get('season_id', type=int) or request.form.get('season_id', type=int)
    season = db.session.get(Season, season_id) if season_id else None
    roster = Roster.query.filter_by(bowler_id=bowler_id, season_id=season_id).first() if season_id else None
    teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all() if season_id else []

    if request.method == 'POST':
        last_name = request.form.get('last_name', '').strip()
        if last_name:
            bowler.last_name = last_name
        bowler.first_name = request.form.get('first_name', '').strip() or None
        bowler.nickname = request.form.get('nickname', '').strip() or None
        bowler.email = request.form.get('email', '').strip() or None

        new_is_editor = 'is_editor' in request.form
        if bowler.is_editor and not new_is_editor:
            editor_count = Bowler.query.filter_by(is_editor=True).count()
            if editor_count <= 1:
                flash('Cannot remove editor status — at least one editor must always exist.', 'danger')
                return redirect(url_for('admin.edit_bowler', bowler_id=bowler_id,
                                        **({'season_id': season_id} if season_id else {})))
        bowler.is_editor = new_is_editor

        if roster:
            roster.team_id = int(request.form['team_id'])
            roster.prior_handicap = int(request.form.get('prior_handicap') or 0)
            roster.joined_week = int(request.form.get('joined_week') or 1)
        db.session.commit()
        flash('Bowler updated.', 'success')
        if season_id:
            return redirect(url_for('admin.season_detail', season_id=season_id))
        return redirect(url_for('admin.seasons'))

    return render_template('admin/edit_bowler.html', bowler=bowler,
                           season=season, roster=roster, teams=teams)


@admin_bp.route('/bowlers/<int:bowler_id>/test-login')
def test_viewer_login(bowler_id):
    """Create a magic link token and redirect straight to it — no email sent.
    Editor-only (enforced by before_request). Use in an incognito window to
    preview the app as a viewer without disturbing your own session."""
    from routes.auth import send_magic_link as _send
    import uuid
    from datetime import datetime, timedelta
    from models import MagicLinkToken

    bowler = Bowler.query.get_or_404(bowler_id)
    if bowler.is_editor:
        flash(f'{bowler.display_name} is an editor — their view is the same as yours. '
              'Pick a non-editor bowler to test viewer mode.', 'warning')
        return redirect(request.referrer or url_for('admin.seasons'))

    now = datetime.utcnow()
    # Invalidate prior unused tokens
    MagicLinkToken.query.filter_by(bowler_id=bowler.id, used_at=None).update({'used_at': now})
    token_str = str(uuid.uuid4())
    db.session.add(MagicLinkToken(
        token=token_str,
        bowler_id=bowler.id,
        expires_at=now + timedelta(hours=1),
        created_at=now,
    ))
    db.session.commit()

    flash(f'Copy this URL into an incognito window to log in as {bowler.display_name} '
          f'(viewer). The link expires in 1 hour.', 'info')
    # Redirect to the token URL so clicking it here logs the editor in as that bowler.
    # Better: show the URL so they can paste it into incognito instead.
    from flask import request as req, url_for as _url
    token_url = _url('auth.validate_token', token=token_str, _external=True)
    return render_template('admin/test_viewer_link.html',
                           bowler=bowler, token_url=token_url)


@admin_bp.route('/seasons/<int:season_id>/roster/<int:roster_id>/toggle', methods=['POST'])
def toggle_active(season_id, roster_id):
    r = Roster.query.get_or_404(roster_id)
    r.active = not r.active
    db.session.commit()
    return redirect(url_for('admin.season_detail', season_id=season_id))


@admin_bp.route('/seasons/<int:season_id>/roster/<int:roster_id>/edit', methods=['GET', 'POST'])
def edit_roster(season_id, roster_id):
    r = Roster.query.get_or_404(roster_id)
    season = Season.query.get_or_404(season_id)
    teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all()

    if request.method == 'POST':
        r.team_id = int(request.form['team_id'])
        r.prior_handicap = int(request.form.get('prior_handicap') or 0)
        r.joined_week = int(request.form.get('joined_week') or 1)
        db.session.commit()
        flash('Roster entry updated.', 'success')
        return redirect(url_for('admin.season_detail', season_id=season_id))

    return render_template('admin/edit_roster.html', r=r, season=season, teams=teams)


# ---------------------------------------------------------------------------
# Schedule management
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/schedule', methods=['GET', 'POST'])
def schedule(season_id):
    season = Season.query.get_or_404(season_id)
    teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all()
    weeks = Week.query.filter_by(season_id=season_id).order_by(Week.week_num).all()
    existing = ScheduleEntry.query.filter_by(season_id=season_id).all()

    # Build lookup: {(week_num, matchup_num): entry}
    sched_map = {(e.week_num, e.matchup_num): e for e in existing}

    # Highest matchup number that actually has data, per week
    week_matchup_counts = {}
    for (wn, mn), e in sched_map.items():
        if e.team1_id or e.team2_id or e.lane_pair:
            week_matchup_counts[wn] = max(week_matchup_counts.get(wn, 0), mn)

    return render_template('admin/schedule.html',
                           season=season, teams=teams,
                           weeks=weeks, sched_map=sched_map,
                           week_matchup_counts=week_matchup_counts)


_INDIV_TOURNAMENT_TYPES = {'indiv_scratch', 'indiv_hcp_1', 'indiv_hcp_2'}


@admin_bp.route('/seasons/<int:season_id>/schedule/save', methods=['POST'])
def save_schedule(season_id):
    """Save schedule entries from form. Expects fields like week_1_matchup_1_t1, etc."""
    season = Season.query.get_or_404(season_id)

    # Build week-type lookup so we can handle solo (individual tournament) weeks
    week_map = {w.week_num: w for w in Week.query.filter_by(season_id=season_id).all()}

    # Collect all fields per matchup first so we never add a partial entry to the
    # session before all fields are set.
    updates = {}
    for key, val in request.form.items():
        # Expected key format: week_{wn}_matchup_{mn}_{field}
        parts = key.split('_')
        if len(parts) < 5 or parts[0] != 'week' or parts[2] != 'matchup':
            continue
        try:
            wn = int(parts[1])
            mn = int(parts[3])
            field = parts[4]
        except (ValueError, IndexError):
            continue
        updates.setdefault((wn, mn), {})[field] = val

    for (wn, mn), fields in updates.items():
        t1 = int(fields['t1']) if fields.get('t1') else None
        t2 = int(fields['t2']) if fields.get('t2') else None
        lane = fields.get('lane', '').strip()

        week = week_map.get(wn)
        is_solo = week and week.tournament_type in _INDIV_TOURNAMENT_TYPES

        entry = ScheduleEntry.query.filter_by(
            season_id=season_id, week_num=wn, matchup_num=mn
        ).first()

        if not entry:
            if is_solo and lane:
                # Individual tournament: create lane-only entry (no teams)
                entry = ScheduleEntry(season_id=season_id, week_num=wn, matchup_num=mn)
                db.session.add(entry)
            elif t1 and t2:
                entry = ScheduleEntry(season_id=season_id, week_num=wn, matchup_num=mn)
                db.session.add(entry)
            else:
                continue

        if is_solo:
            if lane:
                entry.lane_pair = lane
            else:
                # Lane cleared on a solo entry — delete it
                db.session.delete(entry)
                continue
        else:
            # Always write all three fields so clearing "—" takes effect
            entry.team1_id = t1 or None
            entry.team2_id = t2 or None
            entry.lane_pair = lane or None
            if not t1 and not t2 and not lane:
                # Everything cleared — remove the entry entirely
                db.session.delete(entry)
                continue

    # Delete entries beyond the selected matchup count for non-individual weeks
    for key, val in request.form.items():
        parts = key.split('_')
        # key format: week_{wn}_matchup_count
        if len(parts) == 4 and parts[0] == 'week' and parts[2] == 'matchup' and parts[3] == 'count':
            try:
                wn, count = int(parts[1]), int(val)
            except ValueError:
                continue
            week = week_map.get(wn)
            if week and week.tournament_type not in _INDIV_TOURNAMENT_TYPES:
                (ScheduleEntry.query
                 .filter_by(season_id=season_id, week_num=wn)
                 .filter(ScheduleEntry.matchup_num > count)
                 .delete())

    db.session.commit()
    flash('Schedule saved.', 'success')
    return redirect(url_for('admin.schedule', season_id=season_id))


# ---------------------------------------------------------------------------
# Home message
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/home-message', methods=['POST'])
def save_home_message(season_id):
    """Save the optional home-page message (editor only)."""
    from flask_login import current_user
    if not current_user.is_authenticated or not current_user.is_editor:
        abort(403)
    season = Season.query.get_or_404(season_id)
    season.home_message = request.form.get('home_message', '').strip() or None
    db.session.commit()
    return redirect(url_for('index'))


# ---------------------------------------------------------------------------
# Week date editing
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/weeks', methods=['GET', 'POST'])
def edit_weeks(season_id):
    season = Season.query.get_or_404(season_id)
    weeks = Week.query.filter_by(season_id=season_id).order_by(Week.week_num).all()

    labels = season.tournament_labels
    TOURNAMENT_TYPES = [
        ('', '— Regular week —'),
        ('club_championship', labels['club_championship']),
        ('indiv_scratch',     labels['indiv_scratch']),
        ('indiv_hcp_1',       labels['indiv_hcp_1']),
        ('indiv_hcp_2',       labels['indiv_hcp_2']),
    ]

    if request.method == 'POST':
        # Save tournament display names and venue
        season.name_club_championship = request.form.get('name_club_championship', '').strip() or season.name_club_championship
        season.name_indiv_scratch     = request.form.get('name_indiv_scratch', '').strip()     or season.name_indiv_scratch
        season.name_indiv_hcp_1       = request.form.get('name_indiv_hcp_1', '').strip()       or season.name_indiv_hcp_1
        season.name_indiv_hcp_2       = request.form.get('name_indiv_hcp_2', '').strip()       or season.name_indiv_hcp_2
        venue = request.form.get('venue', '').strip()
        if venue in ('mountain_lakes_club', 'boonton_lanes'):
            season.venue = venue

        for wk in weeks:
            date_str = request.form.get(f'date_{wk.week_num}')
            pos_night = request.form.get(f'pos_{wk.week_num}') == 'on'
            tournament_type = request.form.get(f'tournament_{wk.week_num}') or None
            if date_str:
                wk.date = date.fromisoformat(date_str)
            wk.is_position_night = pos_night
            wk.tournament_type = tournament_type
        db.session.commit()
        flash('Week dates saved.', 'success')
        return redirect(url_for('admin.season_detail', season_id=season_id))

    return render_template('admin/edit_weeks.html', season=season, weeks=weeks,
                           tournament_types=TOURNAMENT_TYPES)


# ---------------------------------------------------------------------------
# Matchup assignment (historical data fix)
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/assign_matchups')
def assign_matchups_list(season_id):
    season = Season.query.get_or_404(season_id)
    weeks = (Week.query
             .filter_by(season_id=season_id, is_entered=True, is_cancelled=False)
             .order_by(Week.week_num)
             .all())
    return render_template('admin/assign_matchups_list.html', season=season, weeks=weeks)


@admin_bp.route('/seasons/<int:season_id>/assign_matchups/<int:week_num>',
                methods=['GET', 'POST'])
def assign_matchups(season_id, week_num):
    season = Season.query.get_or_404(season_id)
    week = Week.query.filter_by(season_id=season_id, week_num=week_num).first_or_404()

    schedules = (ScheduleEntry.query
                 .filter_by(season_id=season_id, week_num=week_num)
                 .order_by(ScheduleEntry.matchup_num)
                 .all())

    # Build pairings: matchups 1&2 → pairing A, matchups 3&4 → pairing B
    pairings = []
    for base in [1, 3]:
        s1 = next((s for s in schedules if s.matchup_num == base), None)
        s2 = next((s for s in schedules if s.matchup_num == base + 1), None)
        if s1:
            pairings.append({'mnum1': base, 'mnum2': base + 1,
                             'sched1': s1, 'sched2': s2,
                             'teams': [s1.team1, s1.team2]})

    if request.method == 'POST':
        for key, val in request.form.items():
            if key.startswith('entry_'):
                entry_id = int(key[6:])
                entry = db.session.get(MatchupEntry, entry_id)
                if entry:
                    entry.matchup_num = int(val)

        db.session.flush()

        # Remove old blinds, re-add to balance each matchup
        MatchupEntry.query.filter_by(
            season_id=season_id, week_num=week_num, is_blind=True
        ).delete()
        db.session.flush()

        for pairing in pairings:
            for mnum in [pairing['mnum1'], pairing['mnum2']]:
                sched = pairing['sched1'] if mnum == pairing['mnum1'] else pairing['sched2']
                if not sched:
                    continue
                for team, other_team, side in [
                    (sched.team1, sched.team2, 'A'),
                    (sched.team2, sched.team1, 'B'),
                ]:
                    c_self = MatchupEntry.query.filter_by(
                        season_id=season_id, week_num=week_num,
                        matchup_num=mnum, team_id=team.id, is_blind=False
                    ).count()
                    c_other = MatchupEntry.query.filter_by(
                        season_id=season_id, week_num=week_num,
                        matchup_num=mnum, team_id=other_team.id, is_blind=False
                    ).count()
                    for _ in range(max(0, c_other - c_self)):
                        db.session.add(MatchupEntry(
                            season_id=season_id, week_num=week_num,
                            matchup_num=mnum, team_id=team.id,
                            is_blind=True, lane_side=side,
                            game1=season.blind_scratch,
                            game2=season.blind_scratch,
                            game3=season.blind_scratch,
                        ))

        db.session.commit()
        flash(f'Week {week_num} assignments saved.', 'success')

        # Advance to next unentered week
        next_week = (Week.query
                     .filter_by(season_id=season_id, is_entered=True, is_cancelled=False)
                     .filter(Week.week_num > week_num)
                     .order_by(Week.week_num)
                     .first())
        if next_week:
            return redirect(url_for('admin.assign_matchups',
                                    season_id=season_id, week_num=next_week.week_num))
        return redirect(url_for('admin.assign_matchups_list', season_id=season_id))

    # GET — load non-blind entries per team for each pairing
    pairing_data = []
    for pairing in pairings:
        team_entries = {}
        for team in pairing['teams']:
            entries = (MatchupEntry.query
                       .filter_by(season_id=season_id, week_num=week_num,
                                  team_id=team.id, is_blind=False)
                       .join(Bowler, MatchupEntry.bowler_id == Bowler.id)
                       .order_by(Bowler.last_name)
                       .all())
            team_entries[team.id] = entries
        pairing_data.append({**pairing, 'team_entries': team_entries})

    return render_template('admin/assign_matchups.html',
                           season=season, week=week, pairing_data=pairing_data)


# ---------------------------------------------------------------------------
# XLS season import (web UI)
# ---------------------------------------------------------------------------

NON_BOWLER_SHEETS = {
    'Instructions', 'Parameters', '2025 Banquet', 'wkly alpha', 'YTD alpha',
    'wkly high average', 'High Games ', 'team scoring', 'dummy', 'blinds',
    'Payout Formula', 'indiv payout', 'final handicap',
}

SKIP_NAMES = {'6 games worksheet', 'weekly highs'}


@admin_bp.route('/import_season', methods=['GET', 'POST'])
def import_season():
    if request.method == 'POST':
        try:
            import openpyxl
        except ImportError:
            flash('openpyxl is not installed. Run: pip install openpyxl', 'danger')
            return redirect(url_for('admin.import_season'))

        f = request.files.get('xls_file')
        if not f or not f.filename:
            flash('No file selected.', 'danger')
            return redirect(url_for('admin.import_season'))

        season_name = request.form.get('season_name', '').strip()
        num_weeks   = int(request.form.get('num_weeks', 23))
        half_boundary = int(request.form.get('half_boundary_week', 11))
        start_date_str = request.form.get('start_date', '')

        if not season_name:
            flash('Season name is required.', 'danger')
            return redirect(url_for('admin.import_season'))

        if Season.query.filter_by(name=season_name).first():
            flash(f'Season "{season_name}" already exists.', 'danger')
            return redirect(url_for('admin.import_season'))

        try:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        except Exception as e:
            flash(f'Could not read spreadsheet: {e}', 'danger')
            return redirect(url_for('admin.import_season'))

        # Deactivate current active season
        Season.query.filter_by(is_active=True).update({'is_active': False})

        season = Season(
            name=season_name,
            num_weeks=num_weeks,
            half_boundary_week=half_boundary,
            is_active=True,
            bowling_format='single',
        )
        if start_date_str:
            season.start_date = date.fromisoformat(start_date_str)
        db.session.add(season)
        db.session.flush()

        # Create 4 teams
        teams = []
        for i in range(1, 5):
            tname = request.form.get(f'team{i}_name', '').strip() or f'Team {i}'
            t = Team(season_id=season.id, number=i, name=tname)
            db.session.add(t)
            teams.append(t)
        db.session.flush()
        team_map = {t.number: t for t in teams}

        # Create week records
        for wn in range(1, num_weeks + 1):
            wk = Week(
                season_id=season.id,
                week_num=wn,
                is_position_night=(wn in [half_boundary, num_weeks - 1]),
            )
            if season.start_date:
                wk.date = season.start_date + timedelta(weeks=wn - 1)
            db.session.add(wk)

        # Post-season tournament weeks
        _add_postseason_weeks(season, num_weeks)
        db.session.flush()

        # ── Import roster from 'wkly alpha' sheet ──────────────────────────
        if 'wkly alpha' not in wb.sheetnames:
            flash('Spreadsheet is missing the "wkly alpha" sheet.', 'danger')
            db.session.rollback()
            return redirect(url_for('admin.import_season'))

        ws_alpha = wb['wkly alpha']
        bowler_added = bowler_skipped = 0

        for row in ws_alpha.iter_rows(min_row=8, max_row=70, min_col=1, max_col=21, values_only=True):
            last_name = row[0]
            if not last_name or not isinstance(last_name, str):
                continue
            if last_name.strip().lower() in SKIP_NAMES:
                continue

            first_name = (row[1] or '').strip() or None
            nickname   = (row[2] or '').strip() or None
            team_num   = row[3]
            curr_hcp   = int(row[9]) if row[9] else 0
            active     = str(row[15] or '').strip().lower() == 'yes'
            email      = (row[19] or '').strip() or None

            if team_num not in team_map:
                bowler_skipped += 1
                continue

            bowler = Bowler.query.filter_by(last_name=last_name.strip()).first()
            if not bowler:
                bowler = Bowler(last_name=last_name.strip(), first_name=first_name,
                                nickname=nickname, email=email)
                db.session.add(bowler)
                db.session.flush()
            else:
                if not bowler.nickname and nickname:
                    bowler.nickname = nickname
                if not bowler.email and email:
                    bowler.email = email

            roster_entry = Roster.query.filter_by(
                bowler_id=bowler.id, season_id=season.id
            ).first()
            if not roster_entry:
                roster_entry = Roster(
                    bowler_id=bowler.id, season_id=season.id,
                    team_id=team_map[team_num].id, active=active,
                    prior_handicap=curr_hcp, joined_week=1,
                )
                db.session.add(roster_entry)
            else:
                roster_entry.team_id = team_map[team_num].id
                roster_entry.active = active
                roster_entry.prior_handicap = curr_hcp
            bowler_added += 1

        db.session.flush()

        # ── Import per-bowler scores from individual sheets ─────────────────
        # matchup_num: teams 1&2 → 2 (B-side), teams 3&4 → 4 (B-side)
        def bowler_matchup_num(team_num):
            return 2 if team_num in (1, 2) else 4

        # Build a roster lookup: last_name → (bowler, team_num)
        roster_lookup = {}
        for row in ws_alpha.iter_rows(min_row=8, max_row=70, min_col=1, max_col=5, values_only=True):
            ln = row[0]
            tn = row[3]
            if ln and isinstance(ln, str) and ln.strip().lower() not in SKIP_NAMES and tn in team_map:
                roster_lookup[ln.strip()] = tn

        entries_added = 0
        for sheet_name in wb.sheetnames:
            if sheet_name in NON_BOWLER_SHEETS:
                continue
            ws_b = wb[sheet_name]
            rows = list(ws_b.iter_rows(values_only=True))
            if len(rows) < 11:
                continue

            week_row  = rows[6]   # row index 6 = row 7
            game1_row = rows[8]
            game2_row = rows[9]
            game3_row = rows[10]

            team_num = roster_lookup.get(sheet_name)
            if team_num is None:
                continue

            team_obj = team_map[team_num]
            bowler = Bowler.query.filter_by(last_name=sheet_name).first()
            if not bowler:
                continue

            matchup_num = bowler_matchup_num(team_num)
            side = 'B'

            for col in range(2, len(week_row)):
                wn = week_row[col]
                if not isinstance(wn, (int, float)):
                    continue
                wn = int(wn)
                if wn < 1 or wn > num_weeks:
                    continue

                g1 = game1_row[col] if col < len(game1_row) else None
                g2 = game2_row[col] if col < len(game2_row) else None
                g3 = game3_row[col] if col < len(game3_row) else None

                if g1 is None and g2 is None and g3 is None:
                    continue
                if g1 == 0 and g2 == 0 and g3 == 0:
                    continue

                entry = MatchupEntry(
                    season_id=season.id, week_num=wn,
                    matchup_num=matchup_num, team_id=team_obj.id,
                    bowler_id=bowler.id, is_blind=False, lane_side=side,
                    game1=int(g1) if g1 is not None else None,
                    game2=int(g2) if g2 is not None else None,
                    game3=int(g3) if g3 is not None else None,
                )
                db.session.add(entry)
                entries_added += 1

        # ── Import team points from 'team scoring' sheet ────────────────────
        from models import TeamPoints
        points_added = 0
        if 'team scoring' in wb.sheetnames:
            ws_ts = wb['team scoring']
            ts_rows = list(ws_ts.iter_rows(values_only=True))
            week_nums_entered = set()
            for row in ts_rows[7:]:
                if row[0] is None or not isinstance(row[0], (int, float)):
                    continue
                wn = int(row[0])
                if wn < 1 or wn > num_weeks:
                    continue
                # cols: T1A=2, T1B=3, T1tot=4, T2A=5, T2B=6, T2tot=7,
                #       T3A=8, T3B=9, T3tot=10, T4A=11, T4B=12, T4tot=13
                for ti, base_col in enumerate([2, 5, 8, 11], start=1):
                    if ti not in team_map:
                        continue
                    pts_a = row[base_col] or 0
                    pts_b = row[base_col + 1] or 0
                    mnum_a = 1 if ti in (1, 2) else 3
                    mnum_b = 2 if ti in (1, 2) else 4
                    team_id = team_map[ti].id
                    for mnum, pts in [(mnum_a, pts_a), (mnum_b, pts_b)]:
                        if pts > 0:
                            db.session.add(TeamPoints(
                                season_id=season.id, week_num=wn,
                                matchup_num=mnum, team_id=team_id,
                                points_earned=float(pts),
                            ))
                            points_added += 1
                week_nums_entered.add(wn)

            # Mark entered weeks
            for wn in week_nums_entered:
                wk = Week.query.filter_by(season_id=season.id, week_num=wn).first()
                if wk:
                    wk.is_entered = True

        db.session.commit()
        flash(
            f'Season "{season_name}" imported: {bowler_added} bowlers, '
            f'{entries_added} score entries, {points_added} team-point records.',
            'success'
        )
        return redirect(url_for('admin.season_detail', season_id=season.id))

    return render_template('admin/import_season.html')


# ---------------------------------------------------------------------------
# Mailing list management
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/mailing-list')
def mailing_list(season_id):
    season = Season.query.get_or_404(season_id)
    teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all()
    roster = (Roster.query
              .filter_by(season_id=season_id, active=True)
              .join(Bowler)
              .order_by(Bowler.last_name)
              .all())
    missing_email = [r for r in roster if not r.bowler.email]
    return render_template('admin/mailing_list.html',
                           season=season, teams=teams,
                           roster=roster, missing_email=missing_email)


# ---------------------------------------------------------------------------
# Weekly email compose and send
# ---------------------------------------------------------------------------

def _resolve_captain_emails(teams, season_id):
    """Return list of (team, bowler_or_None, email_or_None) for all teams."""
    result = []
    for team in teams:
        if not team.captain_name:
            result.append((team, None, None))
            continue
        # Try to match captain_name against bowler last_name in this season's roster
        roster_entry = (Roster.query
                        .filter_by(season_id=season_id)
                        .join(Bowler)
                        .filter(Bowler.last_name == team.captain_name)
                        .first())
        if roster_entry and roster_entry.bowler.email:
            result.append((team, roster_entry.bowler, roster_entry.bowler.email))
        else:
            result.append((team, None, None))
    return result


def _get_above_average_bowlers(season_id, week_num, threshold=30):
    """Return bowlers who bowled >= threshold pins above their prior running average."""
    entries = (MatchupEntry.query
               .filter_by(season_id=season_id, week_num=week_num, is_blind=False)
               .filter(MatchupEntry.bowler_id.isnot(None))
               .all())
    from calculations import get_bowler_stats
    results = []
    seen = set()
    for entry in entries:
        if entry.bowler_id in seen:
            continue
        seen.add(entry.bowler_id)
        if entry.game_count == 0:
            continue
        prior = get_bowler_stats(entry.bowler_id, season_id, week_num - 1)
        prior_avg = prior.get('running_avg') or 0
        if prior_avg == 0:
            continue
        games = entry.games_night1  # standard 3-game series only
        best_game = max(games) if games else 0
        diff = best_game - prior_avg
        if diff >= threshold:
            results.append({
                'bowler': entry.bowler,
                'team': entry.team,
                'games': games,
                'handicap': prior.get('display_handicap', 0),
                'best_game': best_game,
                'prior_avg': prior_avg,
                'diff': diff,
            })
    return sorted(results, key=lambda r: r['bowler'].last_name)


@admin_bp.route('/seasons/<int:season_id>/week/<int:week_num>/email', methods=['GET', 'POST'])
def email_compose(season_id, week_num):
    from calculations import get_weekly_prizes, get_team_standings
    from flask import current_app

    season = Season.query.get_or_404(season_id)
    week = Week.query.filter_by(season_id=season_id, week_num=week_num).first_or_404()
    teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all()
    settings = db.session.get(LeagueSettings, 1)
    league_name = (settings.league_name if settings else 'MLC Pirate Bowling League')

    captain_info = _resolve_captain_emails(teams, season_id)
    # TO: all captains that have emails (sender's own team excluded only if they want;
    # for now include all 4 — sender can remove themselves from the form)
    to_captains = [(t, b, e) for t, b, e in captain_info if e]
    missing_captains = [(t, b, e) for t, b, e in captain_info if not e]

    # BCC: active roster
    all_roster = (Roster.query
                  .filter_by(season_id=season_id, active=True)
                  .join(Bowler)
                  .order_by(Bowler.last_name)
                  .all())

    prizes = get_weekly_prizes(season_id, week_num)
    above_avg = _get_above_average_bowlers(season_id, week_num)
    standings = get_team_standings(season_id, through_week=week_num)

    week_date_str = week.date.strftime('%B %d, %Y') if week.date else f'Week {week_num}'
    default_subject = f'{league_name} Standings -- Week {week_num}'

    if request.method == 'POST':
        subject = request.form.get('subject', default_subject).strip()
        body_text = request.form.get('body_text', '').strip()
        bcc_scope = request.form.get('bcc_scope', 'all')
        attach_pdf = request.form.get('attach_pdf') == '1'
        to_emails_raw = request.form.get('to_emails', '').strip()

        # Build TO list
        to_list = [e.strip() for e in to_emails_raw.split(',') if e.strip()]

        # Build BCC list
        if bcc_scope == 'all':
            bcc_roster = all_roster
        else:
            try:
                team_num = int(bcc_scope)
                bcc_roster = [r for r in all_roster if r.team.number == team_num]
            except ValueError:
                bcc_roster = all_roster

        bcc_list = list({r.bowler.email for r in bcc_roster if r.bowler.email})

        test_only = request.form.get('test_only') == '1'
        if test_only:
            my_email = current_app.config.get('GRAPH_SENDER_EMAIL', '')
            to_list = [my_email] if my_email else []
            bcc_list = []
            subject = f'[TEST] {subject}'

        # Build HTML email body
        html_body = _build_email_html(body_text, above_avg, season, week)

        # Build optional PDF attachment
        pdf_min_games = int(request.form.get('pdf_min_games', 9) or 9)
        pdf_top10 = request.form.get('pdf_top10') == '1'
        pdf_bytes = None
        if attach_pdf:
            try:
                pdf_bytes = _generate_prizes_pdf(season_id, week_num,
                                                 min_games=pdf_min_games, top10=pdf_top10)
            except Exception as pdf_err:
                flash(f'PDF generation failed (email sent without attachment): {pdf_err}', 'warning')

        # Send via Microsoft Graph API
        try:
            _send_via_graph(
                app_config=current_app.config,
                subject=subject,
                html_body=html_body,
                to_list=to_list,
                bcc_list=bcc_list,
                pdf_attachment=pdf_bytes,
                pdf_filename=f'Week{week_num}_Standings.pdf',
            )
            if test_only:
                flash(f'Test email sent to {to_list[0] if to_list else "you"}.', 'success')
            else:
                flash(f'Email sent to {len(to_list)} captain(s) with {len(bcc_list)} BCC recipients.', 'success')
            return redirect(url_for('entry.week_entry', season_id=season_id, week_num=week_num))

        except Exception as e:
            flash(f'Email send failed: {e}', 'danger')

    graph_configured = bool(current_app.config.get('GRAPH_CLIENT_ID'))
    return render_template('admin/email_compose.html',
                           season=season, week=week, week_num=week_num,
                           teams=teams, settings=settings,
                           to_captains=to_captains,
                           missing_captains=missing_captains,
                           all_roster=all_roster,
                           prizes=prizes,
                           above_avg=above_avg,
                           standings=standings,
                           default_subject=default_subject,
                           league_name=league_name,
                           mail_configured=graph_configured)


def _send_via_graph(app_config, subject, html_body, to_list, bcc_list,
                    pdf_attachment=None, pdf_filename=None):
    """Send email via Microsoft Graph API using client-credentials OAuth2."""
    import base64
    import json
    import urllib.request
    import urllib.parse
    import msal

    tenant_id    = app_config['GRAPH_TENANT_ID']
    client_id    = app_config['GRAPH_CLIENT_ID']
    client_secret= app_config['GRAPH_CLIENT_SECRET']
    sender_email = app_config['GRAPH_SENDER_EMAIL']

    if not all([tenant_id, client_id, client_secret, sender_email]):
        raise RuntimeError('Microsoft Graph credentials not configured. '
                           'Set GRAPH_TENANT_ID, GRAPH_CLIENT_ID, '
                           'GRAPH_CLIENT_SECRET, GRAPH_SENDER_EMAIL in .env')

    # Acquire access token
    authority = f'https://login.microsoftonline.com/{tenant_id}'
    app = msal.ConfidentialClientApplication(
        client_id, authority=authority, client_credential=client_secret)
    result = app.acquire_token_for_client(
        scopes=['https://graph.microsoft.com/.default'])

    if 'access_token' not in result:
        raise RuntimeError(f"Token acquisition failed: {result.get('error_description', result)}")

    token = result['access_token']

    # Build message payload
    to_recipients  = [{'emailAddress': {'address': e}} for e in to_list]
    bcc_recipients = [{'emailAddress': {'address': e}} for e in bcc_list]

    message = {
        'subject': subject,
        'body': {'contentType': 'HTML', 'content': html_body},
        'toRecipients': to_recipients,
        'bccRecipients': bcc_recipients,
    }

    if pdf_attachment and pdf_filename:
        message['attachments'] = [{
            '@odata.type': '#microsoft.graph.fileAttachment',
            'name': pdf_filename,
            'contentType': 'application/pdf',
            'contentBytes': base64.b64encode(pdf_attachment).decode('utf-8'),
        }]

    payload = json.dumps({'message': message, 'saveToSentItems': True}).encode('utf-8')

    url = f'https://graph.microsoft.com/v1.0/users/{sender_email}/sendMail'
    req = urllib.request.Request(
        url,
        data=payload,
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
        },
        method='POST',
    )

    try:
        with urllib.request.urlopen(req) as resp:
            if resp.status not in (200, 202):
                raise RuntimeError(f'Graph API returned {resp.status}')
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')
        raise RuntimeError(f'Graph API error {e.code}: {body}')


def _build_email_html(body_text, above_avg, season, week):
    """Build the HTML email body from user narrative + auto-generated data."""
    import html as h

    above_html = ''
    if above_avg:
        # Group by team, preserving team number order
        teams_seen = {}
        for r in above_avg:
            tid = r['team'].id
            if tid not in teams_seen:
                teams_seen[tid] = {'team': r['team'], 'bowlers': []}
            teams_seen[tid]['bowlers'].append(r)
        groups = sorted(teams_seen.values(), key=lambda g: g['team'].number)

        block = ''
        for g in groups:
            block += f"<br><strong>{h.escape(g['team'].name)}:</strong><br>"
            for r in g['bowlers']:
                b = r['bowler']
                full = h.escape(f"{b.first_name} {b.last_name}" if b.first_name else b.last_name)
                nick = f" - {h.escape(b.nickname)}" if b.nickname else ""
                scores = '/'.join(str(s) for s in r['games'])
                block += f"&nbsp;&nbsp;{full}{nick} - {r['diff']} ({scores}-{r['handicap']})<br>"
        above_html = f'<p>Notable bowling (30+ above average):<br>{block}</p>'

    body_html = body_text.replace('\n', '<br>\n') if body_text else ''

    return f'''<html><body style="font-family:Arial,sans-serif;font-size:14px">
{body_html}
{above_html}
</body></html>'''


def _generate_prizes_pdf(season_id, week_num, min_games=9, top10=False):
    """Render the prizes/standings page to PDF bytes via WeasyPrint."""
    from weasyprint import HTML
    from flask import current_app, render_template as rt
    from calculations import (get_weekly_prizes, get_bowler_stats, get_team_standings,
                               calculate_handicap)
    from models import MatchupEntry, Roster, Bowler

    season = db.session.get(Season, season_id)
    week = Week.query.filter_by(season_id=season_id, week_num=week_num).first()
    prizes = get_weekly_prizes(season_id, week_num)

    all_entries = MatchupEntry.query.filter_by(season_id=season_id, week_num=week_num).all()
    total_wood = sum(
        e.total_pins + (
            (season.blind_handicap if e.is_blind else calculate_handicap(e.bowler_id, season_id, week_num))
            * e.game_count
        )
        for e in all_entries
    )
    player_count = sum(1 for e in all_entries if not e.is_blind)
    blind_games  = sum(e.game_count for e in all_entries if e.is_blind)

    roster_entries = (Roster.query
                      .filter_by(season_id=season_id, active=True)
                      .join(Bowler).order_by(Bowler.last_name).all())
    leaders = []
    for r in roster_entries:
        stats = get_bowler_stats(r.bowler_id, season_id, week_num)
        if stats['cumulative_games'] == 0:
            continue
        leaders.append({
            'bowler': r.bowler, 'team': r.team,
            'average':             stats['running_avg'],
            'games':               stats['cumulative_games'],
            'handicap':            stats['display_handicap'],
            'high_game_scratch':   stats['ytd_high_game_scratch'],
            'high_game_hcp':       stats['ytd_high_game_hcp'],
            'high_series_scratch': stats['ytd_high_series_scratch'],
            'high_series_hcp':     stats['ytd_high_series_hcp'],
        })

    avg_rows = sorted([l for l in leaders if l['games'] >= min_games],
                      key=lambda x: (-x['average'], x['bowler'].last_name))
    if top10:
        top10_hcps = set(sorted({r['handicap'] for r in avg_rows})[:10])
        avg_rows = [r for r in avg_rows if r['handicap'] in top10_hcps]
    full_year       = sorted(get_team_standings(season_id, through_week=week_num), key=lambda s: s['team'].number)
    fh_list         = get_team_standings(season_id, half=1, through_week=week_num)
    sh_list         = get_team_standings(season_id, half=2, through_week=week_num)
    first_half_map  = {s['team'].id: s['points'] for s in fh_list}
    second_half_map = {s['team'].id: s['points'] for s in sh_list}
    fh_max = max(first_half_map.values(),  default=0)
    sh_max = max(second_half_map.values(), default=0)
    fy_max = max((s['points'] for s in full_year), default=0)

    html_str = rt('reports/week_prizes_pdf.html',
                  season=season, week=week,
                  prizes=prizes, leaders=leaders,
                  standings=full_year,
                  first_half_map=first_half_map, second_half_map=second_half_map,
                  fh_max=fh_max, sh_max=sh_max, fy_max=fy_max,
                  avg_rows=avg_rows, min_games=min_games, top10=top10,
                  total_wood=total_wood, player_count=player_count,
                  blind_games=blind_games)

    return HTML(string=html_str).write_pdf()


# ── Backup & Restore ──────────────────────────────────────────────────────────

@admin_bp.route('/backup')
def backup_restore():
    import os
    from datetime import datetime
    from pathlib import Path
    from flask import current_app
    from config import get_db_path, get_backup_dir

    db_path = get_db_path()
    backup_dir = get_backup_dir()

    try:
        db_size_kb = round(os.path.getsize(db_path) / 1024, 1)
    except OSError:
        db_size_kb = '?'

    raw = sorted(backup_dir.glob('league-*.db'), reverse=True)
    backups = []
    for p in raw:
        try:
            size_kb = round(os.path.getsize(p) / 1024, 1)
            created = datetime.fromtimestamp(p.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
        except OSError:
            size_kb, created = '?', '?'
        backups.append({'filename': p.name, 'size_kb': size_kb, 'created': created})

    return render_template('admin/backup_restore.html',
                           db_path=db_path,
                           db_size_kb=db_size_kb,
                           backup_dir=backup_dir,
                           backups=backups)


@admin_bp.route('/backup/now', methods=['POST'])
def backup_now():
    import shutil
    from datetime import datetime
    from pathlib import Path
    from config import get_db_path, get_backup_dir

    db_path = get_db_path()
    backup_dir = get_backup_dir()
    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    dest = backup_dir / f'league-{stamp}.db'
    try:
        shutil.copy2(db_path, dest)
        flash(f'Backup created: {dest.name}', 'success')
    except Exception as e:
        flash(f'Backup failed: {e}', 'danger')
    return redirect(url_for('admin.backup_restore'))


@admin_bp.route('/backup/restore/<filename>', methods=['POST'])
def restore_backup(filename):
    import shutil
    from datetime import datetime
    from pathlib import Path
    from config import get_db_path, get_backup_dir

    backup_dir = get_backup_dir()
    src = backup_dir / filename
    if not src.exists() or not src.is_file():
        flash('Backup file not found.', 'danger')
        return redirect(url_for('admin.backup_restore'))

    db_path = get_db_path()
    # Save a pre-restore snapshot before overwriting
    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    pre = backup_dir / f'league-{stamp}-pre-restore.db'
    try:
        shutil.copy2(db_path, pre)
        shutil.copy2(src, db_path)
        flash(f'Restored from {filename}. Previous DB saved as {pre.name}.', 'success')
    except Exception as e:
        flash(f'Restore failed: {e}', 'danger')
    return redirect(url_for('admin.backup_restore'))


# ---------------------------------------------------------------------------
# All Bowlers (cross-season admin view)
# ---------------------------------------------------------------------------

@admin_bp.route('/bowlers')
def all_bowlers():
    filter_mode = request.args.get('filter', 'active')  # 'active' or 'all'

    active_season = Season.query.filter_by(is_active=True).first()

    # Determine which bowlers to list
    if filter_mode == 'active' and active_season:
        active_ids = [r.bowler_id for r in
                      Roster.query.filter_by(season_id=active_season.id, active=True).all()]
        bowlers = (Bowler.query
                   .filter(Bowler.id.in_(active_ids))
                   .order_by(Bowler.last_name, Bowler.first_name)
                   .all())
    else:
        filter_mode = 'all'
        rostered_ids = [r.bowler_id for r in
                        Roster.query.with_entities(Roster.bowler_id).distinct().all()]
        bowlers = (Bowler.query
                   .filter(Bowler.id.in_(rostered_ids))
                   .order_by(Bowler.last_name, Bowler.first_name)
                   .all())

    # Full season history per bowler
    roster_map = {}
    for r in (Roster.query
              .join(Season)
              .order_by(Season.name)
              .all()):
        roster_map.setdefault(r.bowler_id, []).append(r)

    # Which bowlers are already on the current season's roster (any active flag)
    current_season_ids = set()
    if active_season:
        current_season_ids = {r.bowler_id for r in
                              Roster.query.filter_by(season_id=active_season.id).all()}

    return render_template('admin/all_bowlers.html',
                           bowlers=bowlers,
                           roster_map=roster_map,
                           active_season=active_season,
                           current_season_ids=current_season_ids,
                           filter_mode=filter_mode)


@admin_bp.route('/bowlers/<int:bowler_id>/send_otp', methods=['POST'])
def send_bowler_otp(bowler_id):
    from routes.auth import send_otp
    filter_mode = request.args.get('filter', 'active')
    bowler = Bowler.query.get_or_404(bowler_id)
    if not bowler.email:
        flash(f'No email on file for {bowler.display_name}.', 'warning')
        return redirect(url_for('admin.all_bowlers', filter=filter_mode))
    ok, err = send_otp(bowler)
    if ok:
        flash(f'Sign-in code sent to {bowler.email}.', 'success')
    else:
        flash(f'Failed to send OTP: {err}', 'danger')
    return redirect(url_for('admin.all_bowlers', filter=filter_mode))


# ---------------------------------------------------------------------------
# Tournament placement (historical results entry: 1st / 2nd / 3rd place)
# ---------------------------------------------------------------------------

@admin_bp.route('/seasons/<int:season_id>/tournament_placement', methods=['GET', 'POST'])
def tournament_placement(season_id):
    season = Season.query.get_or_404(season_id)

    # The 3 individual tournament weeks (not club_championship)
    indiv_weeks = (Week.query
                   .filter(Week.season_id == season_id,
                           Week.tournament_type.in_(['indiv_scratch', 'indiv_hcp_1', 'indiv_hcp_2']))
                   .order_by(Week.week_num)
                   .all())

    # Club championship finalists: first-half and second-half points leaders
    from calculations import get_team_standings
    first_half  = get_team_standings(season_id, half=1)
    second_half = get_team_standings(season_id, half=2)
    fh_winner = first_half[0]['team']  if first_half  else None
    sh_winner = second_half[0]['team'] if second_half else None
    if fh_winner and sh_winner and fh_winner.id == sh_winner.id:
        # Same team won both halves: they play the second-place second-half team
        finalist_teams = [fh_winner]
        if len(second_half) > 1:
            finalist_teams.append(second_half[1]['team'])
    else:
        finalist_teams = [t for t in [fh_winner, sh_winner] if t is not None]
    # Fall back to all season teams if points data isn't available
    if not finalist_teams:
        finalist_teams = Team.query.filter_by(season_id=season_id).order_by(Team.number).all()
    season_teams = finalist_teams

    # All roster bowlers for this season (active + inactive), for the dropdowns
    roster_bowlers = (Bowler.query
                      .join(Roster, Roster.bowler_id == Bowler.id)
                      .filter(Roster.season_id == season_id)
                      .order_by(Bowler.last_name, Bowler.first_name)
                      .all())

    if request.method == 'POST':
        # ── Individual tournament placements ──────────────────────
        PLACE_SCORES = {1: 300, 2: 200, 3: 100}

        for wk in indiv_weeks:
            tt = wk.tournament_type
            TournamentEntry.query.filter_by(season_id=season_id, week_num=wk.week_num).delete()

            has_entry = False
            for place, score in PLACE_SCORES.items():
                bowler_val = request.form.get(f'{tt}_place{place}', '').strip()
                if not bowler_val:
                    continue
                te = TournamentEntry(
                    season_id=season_id,
                    week_num=wk.week_num,
                    bowler_id=int(bowler_val),
                    handicap=0,
                    game1=score,
                    place=place,
                )
                db.session.add(te)
                has_entry = True

            if has_entry and not wk.is_entered:
                wk.is_entered = True

        # ── Club Championship team placements (place 1=first-half winner, 2=second-half winner)
        ClubChampionshipResult.query.filter_by(season_id=season_id).delete()
        for place in range(1, 3):
            team_val = request.form.get(f'club_place{place}', '').strip()
            if not team_val:
                continue
            db.session.add(ClubChampionshipResult(
                season_id=season_id,
                team_id=int(team_val),
                place=place,
            ))

        db.session.commit()
        cache.clear()  # bust Records cache so new placements are visible immediately
        flash('Tournament placements saved.', 'success')
        return redirect(url_for('admin.season_detail', season_id=season_id))

    # GET — pre-populate individual entries (sorted by game1 desc = placement order)
    existing = {}
    for wk in indiv_weeks:
        entries = (TournamentEntry.query
                   .filter_by(season_id=season_id, week_num=wk.week_num)
                   .order_by(TournamentEntry.game1.desc())
                   .all())
        existing[wk.tournament_type] = entries

    # GET — pre-populate club championship results
    club_results = {r.place: r for r in
                    ClubChampionshipResult.query.filter_by(season_id=season_id).all()}

    return render_template('admin/tournament_placement.html',
                           season=season,
                           indiv_weeks=indiv_weeks,
                           roster_bowlers=roster_bowlers,
                           existing=existing,
                           season_teams=season_teams,
                           club_results=club_results,
                           tournament_labels=season.tournament_labels)

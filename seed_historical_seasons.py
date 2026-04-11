"""
Import all 6 historical bowling seasons from XLS spreadsheets.

Usage (Flask app must be stopped first):
    python3 seed_historical_seasons.py [--dry-run] [season_name]

Examples:
    python3 seed_historical_seasons.py              # import all seasons
    python3 seed_historical_seasons.py 2021-2022    # import one season
    python3 seed_historical_seasons.py --dry-run    # validate without writing

Idempotent: skips seasons that already exist in the DB.
"""

import sys, os, warnings, traceback
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from datetime import date, timedelta

DRY_RUN = '--dry-run' in sys.argv
SEASON_FILTER = next((a for a in sys.argv[1:] if not a.startswith('--')), None)

SPREADSHEET_DIR = os.path.expanduser('~/OneDrive - DGLC/Claude/Historic Scoresheets/')

# Sheets that are NOT individual bowler sheets
NON_BOWLER = {
    'Instructions', 'Parameters', 'wkly alpha', 'YTD alpha',
    'wkly high average', 'High Games ', 'High Games', 'team scoring',
    'dummy', 'blinds', 'Payout Formula', 'indiv payout', 'final handicap',
    'Banquet', 'Banquet Pivot', 'Team Counts', 'Club Scoring',
    'Sheet1', 'Prize not paid',  # miscellaneous non-bowler sheets
}

# Season configurations
SEASONS = [
    {
        'filename': 'scoring 2017-2018 - week 23.xlsx',
        'name': '2017-2018',
        'num_weeks': 22,
        'half_boundary_week': 11,
        'name_club_championship': 'Club Championship',
        'name_indiv_scratch':     'Harry E. Russell Championship',
        'name_indiv_hcp_1':       'Buz Bedford Championship',
        'name_indiv_hcp_2':       'Rose Bowl',
    },
    {
        'filename': 'scoring 2018-2019 - week 23.xlsx',
        'name': '2018-2019',
        'num_weeks': 22,
        'half_boundary_week': 11,
        'name_club_championship': 'Club Championship',
        'name_indiv_scratch':     'Harry E. Russell Championship',
        'name_indiv_hcp_1':       'Buz Bedford Championship',
        'name_indiv_hcp_2':       'Rose Bowl',
    },
    {
        'filename': 'scoring 2019-2020 - week 20.xlsx',
        'name': '2019-2020',
        'num_weeks': 22,      # intended length; only 20 weeks of data (COVID)
        'data_weeks': 20,     # how many weeks actually have scores
        'half_boundary_week': 11,
        'name_club_championship': 'Club Championship',
        'name_indiv_scratch':     'Harry E. Russell Championship',
        'name_indiv_hcp_1':       'Buz Bedford Championship',
        'name_indiv_hcp_2':       'Rose Bowl',
        'covid_season': True,
    },
    {
        'filename': 'scoring 2021-2022 - Week 23.xlsx',
        'name': '2021-2022',
        'num_weeks': 22,
        'half_boundary_week': 11,
        'name_club_championship': 'Club Championship',
        'name_indiv_scratch':     'Harry E. Russell Championship',
        'name_indiv_hcp_1':       'Buz Bedford Championship',
        'name_indiv_hcp_2':       'Rose Bowl',
    },
    {
        'filename': 'scoring 2022-2023 - Week 23.xlsx',
        'name': '2022-2023',
        'num_weeks': 22,
        'half_boundary_week': 11,
        'name_club_championship': 'Club Championship',
        'name_indiv_scratch':     'Harry E. Russell Championship',
        'name_indiv_hcp_1':       'Buz Bedford Championship',
        'name_indiv_hcp_2':       'Rose Bowl',
    },
    {
        'filename': 'scoring 2023-2024 - Week 23.xlsx',
        'name': '2023-2024',
        'num_weeks': 22,
        'half_boundary_week': 11,
        'name_club_championship': 'Club Championship',
        'name_indiv_scratch':     'Harry E. Russell Championship',
        'name_indiv_hcp_1':       'Shep Belyea Open',
        'name_indiv_hcp_2':       'Chad Harris Memorial Bowl',
    },
    {
        'filename': 'scoring 2024-2025 - Week 23.xlsx',
        'name': '2024-2025',
        'num_weeks': 22,
        'half_boundary_week': 11,
        'name_club_championship': 'Club Championship',
        'name_indiv_scratch':     'Harry E. Russell Championship',
        'name_indiv_hcp_1':       'Shep Belyea Open',
        'name_indiv_hcp_2':       'Chad Harris Memorial Bowl',
        'team_name_overrides': {2: 'Pinheads'},  # Team 2 renamed this season
    },
]

# ── Spreadsheet helpers ────────────────────────────────────────────────────────

def get_bowler_sheets(wb):
    return [s for s in wb.sheetnames
            if s not in NON_BOWLER and not s.startswith('20')]


def detect_bowling_format(wb):
    """'double' if any bowler has scores in G4-6 slots; otherwise 'single'."""
    for bname in get_bowler_sheets(wb)[:10]:
        ws = wb[bname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 15:
            continue
        # rows[11..13] = Game 4, 5, 6 (0-indexed)
        for row_idx in [11, 12, 13]:
            row = rows[row_idx]
            for col in range(2, 25):
                if col < len(row) and row[col] is not None and isinstance(row[col], (int, float)) and row[col] > 0:
                    return 'double'
    return 'single'


def read_week_dates(wb, num_weeks):
    """Return {week_num: date} from the first available bowler sheet."""
    for bname in get_bowler_sheets(wb):
        ws = wb[bname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 8:
            continue
        # Row index 7 = 'Date' row; cols 2..num_weeks+1 = weeks 1..num_weeks
        date_row = rows[7]
        dates = {}
        for i in range(num_weeks):
            col = 2 + i
            if col < len(date_row) and date_row[col]:
                v = date_row[col]
                if hasattr(v, 'date'):
                    dates[i + 1] = v.date()
                elif isinstance(v, date):
                    dates[i + 1] = v
        if dates:
            return dates
    return {}


def read_team_names_and_points(wb, num_weeks):
    """
    From team scoring sheet, return:
      team_names: {team_num: captain_name}
      team_pts:   {team_num: {week_num: {'wed': pts, 'thur': pts, 'total': pts}}}
    """
    if 'team scoring' not in wb.sheetnames:
        return {}, {}

    ws = wb['team scoring']
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row containing "Wed"/"Thur" or "A"/"B" (format changed in 2024-2025)
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and row[2] in ('Wed', 'wed', 'A', 'a'):
            header_row_idx = i
            break
    if header_row_idx is None:
        return {}, {}

    # Row just before header has team names
    name_row = rows[header_row_idx - 1]
    team_names = {}
    for team_num in range(1, 5):
        col = 2 + (team_num - 1) * 3
        if col < len(name_row) and name_row[col]:
            raw = str(name_row[col])
            # "Team 1 (Lewis)" → captain = "Lewis"
            import re
            m = re.search(r'\((.+?)\)', raw)
            team_names[team_num] = m.group(1) if m else ''

    team_pts = {n: {} for n in range(1, 5)}
    for row in rows[header_row_idx + 1:]:
        if not row or not isinstance(row[0], (int, float)):
            continue
        week_num = int(row[0])
        if week_num < 1 or week_num > num_weeks:
            continue
        for team_num in range(1, 5):
            base = 2 + (team_num - 1) * 3
            wed   = row[base]   if base   < len(row) else None
            thur  = row[base+1] if base+1 < len(row) else None
            total = row[base+2] if base+2 < len(row) else None
            team_pts[team_num][week_num] = {
                'wed':   float(wed)   if isinstance(wed,   (int, float)) else 0.0,
                'thur':  float(thur)  if isinstance(thur,  (int, float)) else 0.0,
                'total': float(total) if isinstance(total, (int, float)) else 0.0,
            }

    return team_names, team_pts


def read_wkly_alpha_roster(wb):
    """
    Return list of dicts: {last, first, team, prior_hcp, active, email}
    Auto-detects column layout from header row.
    """
    ws = wb['wkly alpha']
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (has 'Name' in col 0)
    header_row = None
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == 'Name':
            header_row = row
            header_idx = i
            break
    if header_row is None:
        return []

    # Map column names → indices
    col = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}
    # Handle multi-line header values
    col_map = {}
    for i, v in enumerate(header_row):
        if v is None:
            continue
        key = str(v).strip().replace('\n', ' ')
        col_map[key] = i

    def gcol(names, default=None):
        for n in names:
            if n in col_map:
                return col_map[n]
        return default

    idx_last  = gcol(['Name'], 0)
    idx_first = gcol(['First'], 1)
    idx_team  = gcol(['Team', 'Bname'], None)
    # Distinguish Bname vs Team columns: if 'Bname' exists, Team is next column
    if 'Bname' in col_map:
        idx_team = col_map.get('Team', 3)
    else:
        idx_team = col_map.get('Team', 2)

    idx_prior_hcp = gcol(["Last Year's Handicap", "Last\nYear's\nHandicap"], None)
    # Normalise multiline
    for k, v in list(col_map.items()):
        if 'Last' in k and 'Year' in k and 'Handicap' in k:
            idx_prior_hcp = v
            break

    idx_active = gcol(['Active'], None)
    idx_email  = gcol(['Email'], None)

    bowlers = []
    for row in rows[header_idx + 1:]:
        if not row or not row[idx_last]:
            continue
        last = str(row[idx_last]).strip()
        if not last or last.lower() in ('total', 'name', 'ave', 'average'):
            continue
        first = str(row[idx_first]).strip() if idx_first is not None and row[idx_first] else ''
        team_val = row[idx_team] if idx_team is not None and idx_team < len(row) else None
        team = int(team_val) if isinstance(team_val, (int, float)) and team_val else None

        prior_hcp_val = row[idx_prior_hcp] if idx_prior_hcp is not None and idx_prior_hcp < len(row) else None
        prior_hcp = int(prior_hcp_val) if isinstance(prior_hcp_val, (int, float)) else 0

        active_val = row[idx_active] if idx_active is not None and idx_active < len(row) else 'Yes'
        active = str(active_val).strip().lower() not in ('no', '0', 'false', '') if active_val is not None else True

        email_val = row[idx_email] if idx_email is not None and idx_email < len(row) else None
        email = str(email_val).strip() if email_val else ''

        bowlers.append({
            'last': last, 'first': first, 'team': team,
            'prior_hcp': prior_hcp, 'active': active, 'email': email,
        })

    return bowlers


def read_bowler_sheet_scores(ws, num_weeks):
    """
    Read one individual bowler sheet.
    Returns: {
        'last': str, 'first': str, 'team': int, 'active': bool,
        'prior_hcp': int,
        'scores': {week_num: [g1,g2,g3,g4,g5,g6]}  # None for missing games
    }
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 10:
        return None

    # Row 2 (index 1): last_name at col 5, first at col 6, team at col 7, active at col 8
    info_row = rows[1]
    last  = str(info_row[5]).strip() if info_row[5] else ''
    first = str(info_row[6]).strip() if len(info_row) > 6 and info_row[6] else ''
    team_val = info_row[7] if len(info_row) > 7 else None
    team = int(team_val) if isinstance(team_val, (int, float)) else None
    active_val = info_row[8] if len(info_row) > 8 else None
    # Active can be 0/1 or 'Yes'/'No'
    if isinstance(active_val, (int, float)):
        active = bool(active_val)
    elif isinstance(active_val, str):
        active = active_val.strip().lower() not in ('no', '0', 'false', '')
    else:
        active = True  # assume active if not specified

    # Row 5 (index 4): prior year handicap at col 8
    prior_row = rows[4] if len(rows) > 4 else []
    prior_hcp_val = prior_row[8] if len(prior_row) > 8 else None
    prior_hcp = int(prior_hcp_val) if isinstance(prior_hcp_val, (int, float)) else 0

    # Rows 9-14 (index 8-13): Game 1-6; cols 2..num_weeks+1 = weeks 1..num_weeks
    game_rows = [rows[8+g] if 8+g < len(rows) else [] for g in range(6)]

    scores = {}
    for wk in range(1, num_weeks + 1):
        col = 1 + wk  # col 2 = week 1, col 3 = week 2, ...
        games = []
        for g in range(6):
            row = game_rows[g]
            val = row[col] if col < len(row) else None
            v = int(val) if isinstance(val, (int, float)) and val > 0 else None
            games.append(v)
        # Only store weeks where at least one game has a score
        if any(g is not None for g in games):
            scores[wk] = games

    return {
        'last': last, 'first': first, 'team': team, 'active': active,
        'prior_hcp': prior_hcp, 'scores': scores,
    }


def read_payout_winners(wb):
    """
    From Payout Formula sheet, extract tournament placement winners.
    Returns: {tournament_key: {1: name, 2: name, 3: name}}
    tournament_key: 'indiv_scratch', 'indiv_hcp_1', 'indiv_hcp_2'
    """
    if 'Payout Formula' not in wb.sheetnames:
        return {}

    ws = wb['Payout Formula']
    rows = list(ws.iter_rows(values_only=True))

    results = {}
    current_key = None

    # Map spreadsheet tournament labels → our keys
    # We look at the Tournaments section
    in_tournaments = False
    tournament_order = []  # order: hcp_1, hcp_2, scratch (as they appear)

    for row in rows:
        if not any(v is not None for v in row):
            continue
        # Check for "Tournaments" section header
        if row[0] == 'Tournaments':
            in_tournaments = True
            continue
        if row[0] in ('Sub-Total', 'Weekly Prizes', 'Team Play'):
            in_tournaments = False
            continue

        if not in_tournaments:
            continue

        # Second column names the tournament
        if row[1] is not None and row[2] is None:
            label = str(row[1]).strip().lower()
            if 'bedford' in label or 'belyea' in label:
                current_key = 'indiv_hcp_1'
            elif 'rose' in label or 'chad' in label:
                current_key = 'indiv_hcp_2'
            elif 'club' in label:
                current_key = 'indiv_scratch'
            else:
                current_key = None
            if current_key and current_key not in results:
                results[current_key] = {}

        # Third column is place (1st, 2nd, 3rd)
        elif current_key and row[2] is not None and row[5] is not None:
            place_str = str(row[2]).strip().lower()
            winner = str(row[6]).strip() if row[6] else ''
            if not winner:
                continue
            if '1st' in place_str:
                results[current_key][1] = winner
            elif '2nd' in place_str:
                results[current_key][2] = winner
            elif '3rd' in place_str:
                results[current_key][3] = winner

    return results


# ── Main import ────────────────────────────────────────────────────────────────

def import_season(cfg, app):
    from models import (db, Season, Team, Bowler, Roster, Week,
                        MatchupEntry, TeamPoints, TournamentEntry, ScheduleEntry)

    name = cfg['name']
    filepath = os.path.join(SPREADSHEET_DIR, cfg['filename'])

    if not os.path.exists(filepath):
        print(f'  ERROR: file not found: {filepath}')
        return False

    print(f'\n{"="*60}')
    print(f'  Importing {name}')
    print(f'{"="*60}')

    with app.app_context():
        # Skip if season already exists
        existing = Season.query.filter_by(name=name).first()
        if existing:
            print(f'  SKIP: season {name} already exists (id={existing.id})')
            return True

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        num_weeks = cfg['num_weeks']
        data_weeks = cfg.get('data_weeks', num_weeks)

        bowling_format = detect_bowling_format(wb)
        print(f'  Bowling format: {bowling_format}')

        week_dates = read_week_dates(wb, data_weeks)
        print(f'  Week dates found: {len(week_dates)} (weeks {min(week_dates) if week_dates else "?"} – {max(week_dates) if week_dates else "?"})')

        start_date = week_dates.get(1)
        print(f'  Start date: {start_date}')

        team_names, team_pts = read_team_names_and_points(wb, data_weeks)
        print(f'  Teams from scoring sheet: {team_names}')

        roster_rows = read_wkly_alpha_roster(wb)
        print(f'  Bowlers in wkly alpha: {len(roster_rows)}')

        payout_winners = read_payout_winners(wb)
        print(f'  Tournament winners found: { {k: list(v.keys()) for k, v in payout_winners.items()} }')

        bowler_sheets = get_bowler_sheets(wb)
        print(f'  Individual bowler sheets: {len(bowler_sheets)}')

        if DRY_RUN:
            print('  DRY RUN — no DB writes.')
            return True

        # ── Create season ──────────────────────────────────────────────────
        season = Season(
            name=name,
            start_date=start_date,
            num_weeks=num_weeks,
            half_boundary_week=cfg['half_boundary_week'],
            is_active=False,
            bowling_format=bowling_format,
            name_club_championship=cfg['name_club_championship'],
            name_indiv_scratch=cfg['name_indiv_scratch'],
            name_indiv_hcp_1=cfg['name_indiv_hcp_1'],
            name_indiv_hcp_2=cfg['name_indiv_hcp_2'],
        )
        db.session.add(season)
        db.session.flush()
        print(f'  Created season id={season.id}')

        # ── Create 4 teams ─────────────────────────────────────────────────
        team_name_overrides = cfg.get('team_name_overrides', {})
        teams = {}
        for num in range(1, 5):
            captain = team_names.get(num, '')
            tname = team_name_overrides.get(num, f'Team {num}')
            team = Team(season_id=season.id, number=num,
                        name=tname, captain_name=captain)
            db.session.add(team)
            teams[num] = team
        db.session.flush()

        # ── Create week records ────────────────────────────────────────────
        from routes.admin import _POSTSEASON_WEEKS
        for wn in range(1, num_weeks + 1):
            wk = Week(
                season_id=season.id,
                week_num=wn,
                date=week_dates.get(wn),
                is_position_night=(wn in [cfg['half_boundary_week'], num_weeks]),
                is_entered=(wn <= data_weeks),
            )
            db.session.add(wk)
        # Post-season tournament weeks
        last_date = week_dates.get(data_weeks)
        for offset, (tt, is_pos) in enumerate(_POSTSEASON_WEEKS, start=1):
            wn = num_weeks + offset
            wk = Week(
                season_id=season.id,
                week_num=wn,
                tournament_type=tt,
                is_position_night=is_pos,
                is_entered=False,
            )
            if last_date:
                wk.date = last_date + timedelta(weeks=offset)
            db.session.add(wk)
        db.session.flush()

        # ── Import bowler scores from individual sheets ─────────────────────
        email_map = {r['last'].lower(): r['email'] for r in roster_rows if r.get('email')}
        active_map = {r['last'].lower(): r['active'] for r in roster_rows}

        bowlers_created = 0
        bowlers_matched = 0
        entries_created = 0
        issues = []

        bowler_data_by_last = {}  # last_name → parsed sheet data

        for bname in bowler_sheets:
            try:
                ws = wb[bname]
                data = read_bowler_sheet_scores(ws, data_weeks)
                if not data or not data['last']:
                    continue

                last = data['last']
                first = data['first']
                team_num = data['team']
                active = data['active']
                prior_hcp = data['prior_hcp']
                scores = data['scores']

                # Cross-reference active status from wkly alpha (more reliable)
                if last.lower() in active_map:
                    active = active_map[last.lower()]
                # If no scores at all, mark inactive
                if not scores:
                    active = False

                # Find or create bowler
                bowler = (Bowler.query
                          .filter_by(last_name=last, first_name=first)
                          .first())
                if not bowler:
                    bowler = Bowler.query.filter_by(last_name=last).first()
                    if bowler and bowler.first_name and first and bowler.first_name != first:
                        # Different first name — create new
                        bowler = None

                if bowler:
                    bowlers_matched += 1
                    # Update email if we have it and they don't
                    if not bowler.email and last.lower() in email_map:
                        bowler.email = email_map[last.lower()]
                else:
                    bowler = Bowler(
                        last_name=last,
                        first_name=first,
                        email=email_map.get(last.lower(), ''),
                    )
                    db.session.add(bowler)
                    db.session.flush()
                    bowlers_created += 1

                # Validate team assignment
                if team_num not in (1, 2, 3, 4):
                    issues.append(f'  WARN: {last} has invalid team {team_num}, skipping roster')
                    continue

                # Create roster entry
                roster = Roster(
                    bowler_id=bowler.id,
                    season_id=season.id,
                    team_id=teams[team_num].id,
                    active=active,
                    prior_handicap=prior_hcp,
                    joined_week=1,
                )
                db.session.add(roster)

                # Create MatchupEntry rows
                for wk_num, games in scores.items():
                    if wk_num > data_weeks:
                        continue
                    g1, g2, g3, g4, g5, g6 = games
                    entry = MatchupEntry(
                        season_id=season.id,
                        week_num=wk_num,
                        matchup_num=team_num,  # simplified historical assignment
                        team_id=teams[team_num].id,
                        bowler_id=bowler.id,
                        is_blind=False,
                        lane_side='A',
                        game1=g1, game2=g2, game3=g3,
                        game4=g4, game5=g5, game6=g6,
                    )
                    db.session.add(entry)
                    entries_created += 1

                bowler_data_by_last[last.lower()] = data

            except Exception as e:
                issues.append(f'  ERROR in sheet {bname}: {e}')
                if '--verbose' in sys.argv:
                    traceback.print_exc()

        db.session.flush()
        print(f'  Bowlers created: {bowlers_created}, matched existing: {bowlers_matched}')
        print(f'  MatchupEntry rows: {entries_created}')

        # ── Team points from scoring sheet ─────────────────────────────────
        pts_created = 0
        for team_num, week_data in team_pts.items():
            for wk_num, pts in week_data.items():
                if wk_num > data_weeks:
                    continue
                total = pts['total']
                if total == 0 and pts['wed'] == 0 and pts['thur'] == 0:
                    continue
                tp = TeamPoints(
                    season_id=season.id,
                    week_num=wk_num,
                    matchup_num=team_num,
                    team_id=teams[team_num].id,
                    points_earned=total,
                )
                db.session.add(tp)
                pts_created += 1
        db.session.flush()
        print(f'  TeamPoints rows: {pts_created}')

        # ── Tournament winners (where known) ──────────────────────────────
        covid = cfg.get('covid_season', False)
        winners_seeded = 0
        if not covid and payout_winners:
            # Map tournament type → post-season week number
            tt_to_wk = {
                'indiv_scratch': num_weeks + 2,
                'indiv_hcp_1':   num_weeks + 3,
                'indiv_hcp_2':   num_weeks + 4,
            }
            for tt, places in payout_winners.items():
                wk_num = tt_to_wk.get(tt)
                if not wk_num:
                    continue
                for place, raw_name in places.items():
                    if not raw_name:
                        continue
                    # Only seed 1st place for now (highest confidence)
                    if place != 1:
                        continue
                    # Try to find bowler by last name
                    last = raw_name.split(',')[0].strip().split()[0]
                    bowler = Bowler.query.filter_by(last_name=last).first()
                    te = TournamentEntry(
                        season_id=season.id,
                        week_num=wk_num,
                        bowler_id=bowler.id if bowler else None,
                        guest_name=None if bowler else raw_name,
                        handicap=0,
                    )
                    db.session.add(te)
                    winners_seeded += 1

        if winners_seeded:
            print(f'  Tournament 1st-place winners seeded: {winners_seeded}')
        elif not covid:
            print(f'  Tournament winners: none available (will need manual entry)')

        db.session.commit()
        print(f'  Season {name} committed successfully.')

        if issues:
            print(f'\n  Issues ({len(issues)}):')
            for msg in issues:
                print(msg)

    return True


def main():
    from app import create_app
    app = create_app()

    seasons_to_import = SEASONS
    if SEASON_FILTER:
        seasons_to_import = [s for s in SEASONS if s['name'] == SEASON_FILTER]
        if not seasons_to_import:
            print(f'ERROR: season {SEASON_FILTER!r} not found. Available: {[s["name"] for s in SEASONS]}')
            sys.exit(1)

    if DRY_RUN:
        print('DRY RUN mode — validating structure only, no DB writes.')

    ok = 0
    fail = 0
    for cfg in seasons_to_import:
        try:
            success = import_season(cfg, app)
            if success:
                ok += 1
            else:
                fail += 1
        except Exception as e:
            print(f'\nFATAL ERROR importing {cfg["name"]}: {e}')
            traceback.print_exc()
            fail += 1

    print(f'\n{"="*60}')
    print(f'Done. {ok} succeeded, {fail} failed.')
    if DRY_RUN:
        print('(DRY RUN — nothing was written)')


if __name__ == '__main__':
    main()

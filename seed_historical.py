#!/usr/bin/env python3
"""
Seed the 2025-2026 historical season from the Week 22 scoring spreadsheet.

This creates the season, teams, roster entries, all 21 completed weeks of
individual scores (MatchupEntry), and team points (TeamPoints) from the
'team scoring' sheet.  Week 22 (position night not yet bowled) is created
as a week record but left unentered.

Run on Mac with Flask app stopped:
  python seed_historical.py "/path/to/scoring 2025-2026 - Week 22.xlsx"

The 2026-2027 season (populated by seed_from_xls.py) can already be in the
DB — bowler records are shared and will be reused.
"""

import sys
import os
from datetime import date

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from app import create_app
from models import db, Season, Team, Bowler, Roster, Week, MatchupEntry, TeamPoints

# ─── constants ────────────────────────────────────────────────────────────────

SEASON_NAME = "2025-2026"
POSITION_NIGHT_WEEKS = {11, 22}
TOTAL_WEEKS = 22

# Sheets that are not individual bowler score sheets
NON_BOWLER_SHEETS = {
    'Instructions', 'Parameters', '2025 Banquet', 'wkly alpha', 'YTD alpha',
    'wkly high average', 'High Games ', 'team scoring', 'dummy', 'blinds',
    'Payout Formula', 'indiv payout', 'final handicap',
}

# ─── helpers ──────────────────────────────────────────────────────────────────

def to_date(v):
    """Convert openpyxl datetime or date to Python date."""
    if v is None:
        return None
    if hasattr(v, 'date'):
        return v.date()
    return v


def load_wkly_alpha(wb):
    """
    Parse the 'wkly alpha' sheet.
    Returns dict: last_name -> {first, nickname, team_num, prior_handicap, active}
    """
    ws = wb['wkly alpha']
    rows = list(ws.iter_rows(values_only=True))
    # Header row at index 6; data rows start at index 7
    # Cols: 0=last, 1=first, 2=nickname, 3=team, 4=total, 5=2nd, 6=games,
    #       7=ave, 8=use_hcp, 9=curr_hcp, 10=HGscr, 11=HGhcp, 12=HSscr,
    #       13=HShcp, 14=prior_year_hcp, 15=active
    result = {}
    for row in rows[7:]:
        last = row[0]
        if last is None:
            continue
        last = str(last).strip()
        if last in ('6 games worksheet', 'weekly highs') or last == '':
            continue
        result[last] = {
            'first':          str(row[1]).strip() if row[1] else '',
            'nickname':       str(row[2]).strip() if row[2] else None,
            'team_num':       int(row[3]) if row[3] else None,
            'prior_handicap': int(row[14]) if row[14] else 0,
            'active':         str(row[15]).strip().lower() == 'yes' if row[15] else False,
        }
    return result


def load_week_dates(wb):
    """
    Parse week dates from the 'blinds' sheet.
    Returns dict: week_num (int) -> date
    """
    ws = wb['blinds']
    rows = list(ws.iter_rows(values_only=True))
    week_row = rows[6]   # row index 6: (1, 'Week', 1, 2, 3, …)
    date_row = rows[7]   # row index 7: (2, 'Date', datetime, …)
    result = {}
    for col in range(2, 25):   # cols C through X
        wk = week_row[col] if col < len(week_row) else None
        dt = date_row[col] if col < len(date_row) else None
        if wk is None or dt is None:
            continue
        result[int(wk)] = to_date(dt)
    return result


def load_bowler_sheet(wb, sheet_name):
    """
    Parse an individual bowler score sheet.
    Returns dict: week_num -> (g1, g2, g3)  — only weeks where games were bowled.
    Week 22 (position night not yet played) is excluded.
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    week_row  = rows[6]   # (1, 'Week', 1, 2, … 22, 23, None)
    game1_row = rows[8]   # (3, 'Game 1', g, g, …)
    game2_row = rows[9]
    game3_row = rows[10]

    scores = {}
    for col in range(2, 25):
        wk = week_row[col] if col < len(week_row) else None
        if wk is None:
            continue
        week_num = int(wk)
        if week_num == 22:          # position night not yet bowled
            continue

        g1 = game1_row[col] if col < len(game1_row) else None
        g2 = game2_row[col] if col < len(game2_row) else None
        g3 = game3_row[col] if col < len(game3_row) else None

        # Skip weeks where the bowler was absent (all None) or all zeros
        if g1 is None and g2 is None and g3 is None:
            continue
        if g1 == 0 and g2 == 0 and g3 == 0:
            continue

        scores[week_num] = (
            int(g1) if g1 is not None else None,
            int(g2) if g2 is not None else None,
            int(g3) if g3 is not None else None,
        )
    return scores


def load_team_points(wb):
    """
    Parse the 'team scoring' sheet.
    Returns list of dicts with week_num and per-team A/B sub-scores.

    Layout (after headers):
      col 0=week, col1=date, col2=T1A, col3=T1B, col4=T1Tot,
      col5=T2A, col6=T2B, col7=T2Tot, col8=T3A, col9=T3B, col10=T3Tot,
      col11=T4A, col12=T4B, col13=T4Tot, col14=grand_total
    """
    ws = wb['team scoring']
    rows = list(ws.iter_rows(values_only=True))

    result = []
    for row in rows[7:]:
        wk = row[0]
        if wk is None or not isinstance(wk, (int, float)):
            continue
        week_num = int(wk)

        def pts(v):
            return float(v) if v is not None else 0.0

        result.append({
            'week_num': week_num,
            't1A': pts(row[2]),
            't1B': pts(row[3]),
            't2A': pts(row[5]),
            't2B': pts(row[6]),
            't3A': pts(row[8]),
            't3B': pts(row[9]),
            't4A': pts(row[11]),
            't4B': pts(row[12]),
        })
    return result


def load_team_names(wb):
    """
    Parse team names from 'team scoring' row 5.
    Returns dict: team_number (1-4) -> short_name  e.g. {1: 'Lewis', 2: 'Ferrante', …}
    """
    ws = wb['team scoring']
    rows = list(ws.iter_rows(values_only=True))
    # Row index 5: (None, None, None, None, None, None, 'Team 1 (Lewis)', …)
    # Team names appear at cols 6, 9, 12, 15 — but let's parse dynamically
    name_row = rows[5]
    teams = {}
    for col, val in enumerate(name_row):
        if val and isinstance(val, str) and val.startswith('Team'):
            # "Team 1 (Lewis)" → number=1, name="Lewis"
            try:
                parts = val.split('(')
                num_part = parts[0].strip()   # "Team 1"
                name_part = parts[1].rstrip(')')  # "Lewis"
                num = int(num_part.split()[1])
                teams[num] = name_part.strip()
            except (IndexError, ValueError):
                pass
    return teams


# ─── main import ──────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python seed_historical.py <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # ── Parse spreadsheet data ────────────────────────────────────────────────
    alpha_data  = load_wkly_alpha(wb)
    week_dates  = load_week_dates(wb)
    team_names  = load_team_names(wb)
    all_tp      = load_team_points(wb)

    bowler_sheets = [s for s in wb.sheetnames if s not in NON_BOWLER_SHEETS]
    print(f"Found {len(bowler_sheets)} bowler sheets, {len(alpha_data)} alpha rows, "
          f"{len(week_dates)} weeks, {len(all_tp)} team-point rows")

    app = create_app()
    with app.app_context():

        # ── Guard: skip if season already exists ──────────────────────────────
        existing = Season.query.filter_by(name=SEASON_NAME).first()
        if existing:
            print(f"\nSeason '{SEASON_NAME}' already exists (id={existing.id}).")
            resp = input("Delete all data for this season and re-import? [y/N] ").strip().lower()
            if resp != 'y':
                print("Aborted.")
                return

            print("Deleting existing data…")
            sid = existing.id
            TeamPoints.query.filter_by(season_id=sid).delete()
            MatchupEntry.query.filter_by(season_id=sid).delete()
            Roster.query.filter_by(season_id=sid).delete()
            Week.query.filter_by(season_id=sid).delete()
            Team.query.filter_by(season_id=sid).delete()
            db.session.delete(existing)
            db.session.commit()
            print("Deleted.")

        # ── 1. Season ─────────────────────────────────────────────────────────
        start = week_dates.get(1)
        season = Season(
            name=SEASON_NAME,
            start_date=start,
            num_weeks=TOTAL_WEEKS,
            half_boundary_week=11,
            handicap_base=200,
            handicap_factor=0.9,
            blind_scratch=125,
            blind_handicap=60,
            is_active=False,
        )
        db.session.add(season)
        db.session.flush()   # get season.id
        print(f"\nCreated season '{SEASON_NAME}' id={season.id}")

        # ── 2. Teams ──────────────────────────────────────────────────────────
        team_map = {}   # number -> Team object
        for num in sorted(team_names):
            t = Team(season_id=season.id, number=num, name=team_names[num])
            db.session.add(t)
            db.session.flush()
            team_map[num] = t
            print(f"  Team {num}: {team_names[num]} (id={t.id})")

        # ── 3. Bowlers & Roster ───────────────────────────────────────────────
        bowler_map = {}   # last_name -> Bowler object
        skipped_alpha = []

        for last_name, info in alpha_data.items():
            team_num = info['team_num']
            if team_num not in team_map:
                skipped_alpha.append(f"{last_name} (no team {team_num})")
                continue

            # Reuse existing Bowler record if present (shared with 2026-2027)
            bowler = Bowler.query.filter_by(last_name=last_name).first()
            if not bowler:
                bowler = Bowler(
                    last_name=last_name,
                    first_name=info['first'],
                    nickname=info['nickname'],
                )
                db.session.add(bowler)
                db.session.flush()

            bowler_map[last_name] = bowler

            roster = Roster(
                bowler_id=bowler.id,
                season_id=season.id,
                team_id=team_map[team_num].id,
                active=info['active'],
                prior_handicap=info['prior_handicap'],
                joined_week=1,
            )
            db.session.add(roster)

        db.session.flush()
        print(f"\nCreated/linked {len(bowler_map)} bowlers")
        if skipped_alpha:
            print(f"  Skipped (no valid team): {skipped_alpha}")

        # ── 4. Weeks ──────────────────────────────────────────────────────────
        week_map = {}   # week_num -> Week object
        for wk_num in range(1, TOTAL_WEEKS + 1):
            dt = week_dates.get(wk_num)
            is_pos = wk_num in POSITION_NIGHT_WEEKS
            is_entered = wk_num < 22    # week 22 position night not yet bowled
            w = Week(
                season_id=season.id,
                week_num=wk_num,
                date=dt,
                is_position_night=is_pos,
                is_entered=is_entered,
            )
            db.session.add(w)
            db.session.flush()
            week_map[wk_num] = w
        print(f"Created {len(week_map)} week records")

        # ── 5. Individual scores (MatchupEntry) ───────────────────────────────
        # matchup_num follows team number (1-4); stats don't depend on matchup_num
        entry_count = 0
        for sheet_name in bowler_sheets:
            if sheet_name not in bowler_map:
                # Sheet exists but bowler not in alpha (shouldn't happen often)
                continue

            bowler = bowler_map[sheet_name]
            # Get team from roster
            roster_entry = Roster.query.filter_by(
                bowler_id=bowler.id, season_id=season.id
            ).first()
            if not roster_entry:
                continue

            team = roster_entry.team
            matchup_num = team.number   # simplified: matchup_num = team number

            try:
                scores = load_bowler_sheet(wb, sheet_name)
            except Exception as e:
                print(f"  Warning: could not read sheet '{sheet_name}': {e}")
                continue

            for week_num, (g1, g2, g3) in scores.items():
                me = MatchupEntry(
                    season_id=season.id,
                    week_num=week_num,
                    matchup_num=matchup_num,
                    team_id=team.id,
                    bowler_id=bowler.id,
                    is_blind=False,
                    lane_side='A',  # lane side not tracked in spreadsheet
                    game1=g1,
                    game2=g2,
                    game3=g3,
                )
                db.session.add(me)
                entry_count += 1

        db.session.flush()
        print(f"Created {entry_count} matchup entries (individual game scores)")

        # ── 6. Team points (TeamPoints) ───────────────────────────────────────
        # Teams 1&2 compete: matchup_nums 1 (A sheet) and 2 (B sheet)
        # Teams 3&4 compete: matchup_nums 3 (A sheet) and 4 (B sheet)
        tp_count = 0
        for row in all_tp:
            wk = row['week_num']

            entries = [
                # (team_num, matchup_num, points)
                (1, 1, row['t1A']),
                (1, 2, row['t1B']),
                (2, 1, row['t2A']),
                (2, 2, row['t2B']),
                (3, 3, row['t3A']),
                (3, 4, row['t3B']),
                (4, 3, row['t4A']),
                (4, 4, row['t4B']),
            ]
            for team_num, matchup_num, pts in entries:
                tp = TeamPoints(
                    season_id=season.id,
                    week_num=wk,
                    matchup_num=matchup_num,
                    team_id=team_map[team_num].id,
                    points_earned=pts,
                )
                db.session.add(tp)
                tp_count += 1

        db.session.flush()
        print(f"Created {tp_count} team-point records")

        # ── Commit ────────────────────────────────────────────────────────────
        db.session.commit()
        print(f"\n✓ Historical season '{SEASON_NAME}' fully imported.")
        print(f"  Weeks 1-21 marked as entered; week 22 (position night) left open.")
        print(f"  Season is set inactive — activate it in Admin when needed.")


if __name__ == '__main__':
    main()
